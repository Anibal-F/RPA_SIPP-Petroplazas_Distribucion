import json
import openpyxl
from pathlib import Path
from typing import List, Tuple

FOLIO_COL           = 4    # Column D  — folio de factura
FECHA_FACTURA_COL   = 5    # Column E  — Fecha Factura (dd/mm/yyyy)
CC_COL              = 32   # Column AF — CC OC
OBS_COL             = 33   # Column AG — Observaciones OC
SUBTOTAL_COL        = 34   # Column AH — Subtotal OC
DESCUENTO_COL       = 35   # Column AI — Descuento OC
IVA_COL             = 36   # Column AJ — IVA (16%) OC
GASTOS_ENVIO_COL    = 37   # Column AK — Gastos de Envío OC
TOTAL_OC_COL        = 38   # Column AL — Total OC
FOLIO_FISCAL_COL    = 39   # Column AM — Folio Fiscal (UUID del CFDI)
CUENTA_CONTABLE_START_COL = 40   # Column AN — primera Cuenta Contable (dinámica)
POLIZA_SIPP_COL     = 70   # Columna fija — JSON con la póliza real (cuenta/cargo/abono)
HEADER_ROW     = 8
DATA_START_ROW = 9


class ExcelHandler:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath)
        self.ws = self.wb.active

    def get_folios(self) -> List[Tuple[int, str, str]]:
        """
        Return list of (row_number, folio, fecha_factura) from columns D/E
        starting at row 9. fecha_factura es el texto tal cual ('dd/mm/yyyy'),
        usado para desambiguar folios duplicados en SIPP por periodo.
        """
        result = []
        for row in range(DATA_START_ROW, self.ws.max_row + 1):
            val = self.ws.cell(row=row, column=FOLIO_COL).value
            if val is not None and str(val).strip():
                fecha = self.ws.cell(row=row, column=FECHA_FACTURA_COL).value
                fecha_str = str(fecha).strip() if fecha is not None else ""
                result.append((row, str(val).strip(), fecha_str))
        return result

    def ensure_headers(self):
        """Write column headers to row 8 if not already present."""
        headers = {
            CC_COL:           "CC OC",
            OBS_COL:          "Observaciones OC",
            SUBTOTAL_COL:     "Subtotal OC",
            DESCUENTO_COL:    "Descuento OC",
            IVA_COL:          "IVA OC",
            GASTOS_ENVIO_COL: "Gastos Envío OC",
            TOTAL_OC_COL:     "Total OC",
            FOLIO_FISCAL_COL: "Folio Fiscal",
            POLIZA_SIPP_COL:  "Poliza SIPP (JSON)",
        }
        for col, title in headers.items():
            if not self.ws.cell(row=HEADER_ROW, column=col).value:
                self.ws.cell(row=HEADER_ROW, column=col, value=title)

    def write_result(
        self,
        row: int,
        cc: str,
        observaciones: str,
        subtotal: str = "",
        descuento: str = "",
        iva: str = "",
        gastos_envio: str = "",
        total_oc: str = "",
        folio_fiscal: str = "",
        cuentas_contables: list = None,
        poliza_lineas: list = None,
    ):
        self.ws.cell(row=row, column=CC_COL,           value=cc)
        self.ws.cell(row=row, column=OBS_COL,          value=observaciones)
        self.ws.cell(row=row, column=SUBTOTAL_COL,     value=subtotal)
        self.ws.cell(row=row, column=DESCUENTO_COL,    value=descuento)
        self.ws.cell(row=row, column=IVA_COL,          value=iva)
        self.ws.cell(row=row, column=GASTOS_ENVIO_COL, value=gastos_envio)
        self.ws.cell(row=row, column=TOTAL_OC_COL,     value=total_oc)
        self.ws.cell(row=row, column=FOLIO_FISCAL_COL, value=folio_fiscal)
        for i, code in enumerate(cuentas_contables or []):
            col = CUENTA_CONTABLE_START_COL + i
            # Escribir header dinámico si aún no existe
            if not self.ws.cell(row=HEADER_ROW, column=col).value:
                self.ws.cell(row=HEADER_ROW, column=col, value=f"Cuenta Contable {i + 1}")
            self.ws.cell(row=row, column=col, value=code)
        if poliza_lineas:
            self.ws.cell(row=row, column=POLIZA_SIPP_COL, value=json.dumps(poliza_lineas, ensure_ascii=False))

    def save(self):
        self.wb.save(self.filepath)
        self.wb.close()

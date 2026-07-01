"""
Genera los archivos .txt de pólizas contables (formato de importación
Contpaq) a partir de un reporte *_comparacion.xlsx ya generado por
compare_sucursales.

Se generan 3 pólizas, reflejando el ciclo físico del almacén — son GLOBALES
(un Cargo/Abono por concepto y factura, sin desagregar por estación; el
reparto entre sucursales es solo informativo y vive en la hoja
"Distrib. Calculada" del reporte de comparación):

  Provisión de Almacén
      Si la factura tiene columna "Poliza SIPP (JSON)" (capturada por el RPA
      desde PolizaGrid del modal "Visualizar Factura"), se usan ESAS líneas
      tal cual — son la póliza real que SIPP ya calculó, con retenciones
      reales incluidas. Si no está disponible (reportes generados antes de
      esta actualización del RPA), se usa una aproximación:
          Cargo  201-01-02-1501  Proveedores Tránsito  = Subtotal de factura
          Cargo  121-01-03-0000  IVA 16% por acreditar  = IVA de factura
          Abono  201-01-02-1500  Proveedores            = suma de los 2 cargos
      (esta aproximación NO contempla retenciones)

  Entrada de Almacén
      Cargo  Almacén de Refacciones de la plaza (Cuentas_AlmacénRef.csv)
      Abono  201-01-02-1501  Proveedores Tránsito
             ambos por el monto real de la factura (cargo de la primera línea
             de la póliza SIPP real; si no hay póliza SIPP, Subtotal OC) —
             libera el tránsito al recibir la mercancía en el almacén.

  Salida de Almacén
      Cargo  Cuenta de Gasto Global (Cuenta Contable 1 que resolvió SIPP)
      Abono  Almacén de Refacciones de la plaza
             ambos por el mismo monto real — consumo del almacén.
             (Subtotal OC NO se usa cuando hay póliza SIPP real: para
             facturas DISTRIBUCIÓN/Corporativo puede reflejar el total de
             toda la Orden de Compra, no el de esta factura.)

Formato de líneas P (148 + CRLF) y M (206 + CRLF) tomado de la macro VBA de
referencia usada para pólizas de banco.
"""
import csv
import json
import re
import unicodedata
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Callable

import openpyxl

from .catalog_io import read_catalog

CUENTA_PROVEEDORES = "201-01-02-1500"
CUENTA_TRANSITO = "201-01-02-1501"
CUENTA_IVA_DEFAULT = "121-01-03-0000"

CUENTAS_DIR = Path(__file__).parent.parent / "CuentasContables"
ALMACEN_REF_CSV = CUENTAS_DIR / "Cuentas_AlmacénRef.csv"
PROVEEDORES_CSV = CUENTAS_DIR / "Cuentas_Proveedores.csv"
GASTO_ESTACIONES_CSV = CUENTAS_DIR / "Cuentas_GastoEstaciones.csv"

TIPO_DIARIO = 3


def _normalize(text: str) -> str:
    text = unicodedata.normalize("NFD", str(text or ""))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text.upper().strip()


def _clean_account(cuenta: str) -> str:
    return str(cuenta).strip().replace("-", "").replace(" ", "")


def _clean_concept(text: str, max_len: int = 80) -> str:
    s = re.sub(r"\s+", " ", str(text or "")).strip().upper()
    return s[:max_len]


def _parse_fecha_factura(raw, fallback: date) -> tuple[date, bool]:
    """Parsea 'Fecha Factura' de Datos Originales. Retorna (fecha, uso_fallback)."""
    if isinstance(raw, datetime):
        return raw.date(), False
    if isinstance(raw, date):
        return raw, False
    s = str(raw or "").strip()
    if s:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date(), False
            except ValueError:
                continue
    return fallback, True


def _parse_amount(v) -> float:
    s = str(v or "").strip().replace("$", "").replace(",", "").replace("%", "").replace(" ", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def load_almacen_catalog() -> dict[str, str]:
    """Lee Cuentas_AlmacénRef.csv. Retorna {nombre_plaza_normalizado: cuenta}."""
    catalog: dict[str, str] = {}
    if not ALMACEN_REF_CSV.exists():
        return catalog
    with open(ALMACEN_REF_CSV, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            cuenta = (row.get("Cuenta") or "").strip()
            nombre = (row.get("Almacén") or "").strip()
            if not cuenta or not nombre:
                continue
            key = _normalize(re.sub(r"^SUCURSAL\s*\(|\)$", "", nombre.upper()))
            catalog.setdefault(key, cuenta)
    return catalog


def _parse_poliza_sipp(raw) -> list[dict]:
    """Parsea la columna 'Poliza SIPP (JSON)' → [{"cuenta","cargo","abono"}, ...]."""
    if not raw or not str(raw).strip():
        return []
    try:
        data = json.loads(raw)
    except (TypeError, ValueError):
        return []
    if not isinstance(data, list):
        return []
    lineas = []
    for item in data:
        if not isinstance(item, dict):
            continue
        cuenta = str(item.get("cuenta") or "").strip()
        if not cuenta:
            continue
        lineas.append({
            "cuenta": cuenta,
            "cargo": _parse_amount(item.get("cargo")),
            "abono": _parse_amount(item.get("abono")),
        })
    return lineas


def _match_catalog(nombre: str, catalog: dict[str, str]) -> str | None:
    """Match exacto (normalizado) y, si no hay, substring en cualquier sentido."""
    key = _normalize(nombre)
    if key in catalog:
        return catalog[key]
    for cat_key, cuenta in catalog.items():
        if cat_key and (cat_key in key or key in cat_key):
            return cuenta
    return None


def _match_almacen(sucursal: str, catalog: dict[str, str]) -> str | None:
    return _match_catalog(sucursal, catalog)


def load_proveedores_catalog() -> dict[str, str]:
    """Lee Cuentas_Proveedores.csv. Retorna {nombre_normalizado: cuenta}."""
    catalog: dict[str, str] = {}
    for row in read_catalog(PROVEEDORES_CSV, ["Cuenta", "Nombre"]):
        cuenta = row["Cuenta"].strip()
        nombre = row["Nombre"].strip()
        if not cuenta or not nombre:
            continue
        catalog.setdefault(_normalize(nombre), cuenta)
    return catalog


_STATION_ALIASES = {
    "EJERCITO": "GRIJALVA",
    "CUAHUTEMOC": "CUAUHTEMOC",
    "LA PAZ": "FORJADORES",
}


def load_estaciones_catalog() -> dict[str, str]:
    """Lee Cuentas_GastoEstaciones.csv. Retorna {nombre_estacion_normalizado: 'XX'}
    a partir de las filas encabezado '502-XX-00-0000' → 'Sucursal (Nombre)'."""
    catalog: dict[str, str] = {}
    for row in read_catalog(GASTO_ESTACIONES_CSV, ["Cuenta", "Nombre"]):
        cuenta = row["Cuenta"].strip()
        nombre = row["Nombre"].strip()
        if not cuenta or not nombre:
            continue
        segs = cuenta.split("-")
        if len(segs) != 4 or segs[2] != "00" or segs[3] != "0000":
            continue  # solo nos interesan los encabezados de estación
        key = _normalize(re.sub(r"^SUCURSAL\s*\(|\)$", "", nombre.upper()))
        catalog.setdefault(key, segs[1])
    for alias, canonical in _STATION_ALIASES.items():
        if canonical in catalog and alias not in catalog:
            catalog[alias] = catalog[canonical]
    return catalog


def _build_p_line(fecha_pol: date, tipo_poliza: int, num_pol: int, concepto: str) -> str:
    concepto_limpio = _clean_concept(concepto, 80)
    encabezado = (
        "P  " + fecha_pol.strftime("%Y%m%d") + "    " + str(tipo_poliza)
        + "      " + str(num_pol) + " 1 0          "
    )
    espacios = max(1, 141 - len(encabezado) - len(concepto_limpio))
    return encabezado + concepto_limpio + (" " * espacios) + "11 0 0 " + "\r\n"


def _build_am_ad_lines(folio_fiscal: str) -> str:
    """
    Líneas de asociación de CFDI para Contpaq, tomadas de la macro VBA de
    referencia (macrovba2): después de cada movimiento M se agregan
        AM <UUID>           (Asocia el comprobante fiscal al movimiento)
        AD <UUID>           (Asocia el documento digital)
    para que Contpaq vincule la factura automáticamente. Sin UUID no se emite
    nada (no rompe pólizas de facturas sin folio fiscal capturado).
    """
    uuid = str(folio_fiscal or "").strip().upper()
    if not uuid:
        return ""
    return "AM " + uuid + " \r\n" + "AD " + uuid + " \r\n"


def _build_m_line(cuenta: str, tipo_mov: int, importe: float, concepto: str,
                   referencia: str = "", folio_fiscal: str = "") -> str:
    # Formato macrovba2 (Sepsa) — 206 chars + CRLF
    # "M  "(3) + cuenta(30) + " " + referencia(10) + " " + tipo(1) + " " +
    # importe(20) + " " + diario(10) + " " + importeME(20) + " " + concepto(100) +
    # " " + "    " + " " (6 trailing = VBA: " " & ftSpace(" ",4,"D") & " ")
    # Si la factura trae folio fiscal (UUID), se anexan las líneas AM/AD para
    # que Contpaq asocie el CFDI al movimiento automáticamente.
    return (
        "M  "
        + _clean_account(cuenta).ljust(30)[:30]
        + " " + _clean_concept(referencia, 10).ljust(10)
        + " " + str(tipo_mov)
        + " " + f"{importe:.2f}".ljust(20)[:20]
        + " " + "0".ljust(10)
        + " " + "0.0".ljust(20)
        + " " + _clean_concept(concepto, 100).ljust(100)
        + "      "
        + "\r\n"
        + _build_am_ad_lines(folio_fiscal)
    )


def _write_poliza(path: Path, fecha_poliza: date, tipo_poliza: int, num_poliza: int,
                   concepto_poliza: str, movimientos: str):
    contenido = _build_p_line(fecha_poliza, tipo_poliza, num_poliza, concepto_poliza) + movimientos
    with open(path, "wb") as f:
        f.write(contenido.encode("cp1252", errors="replace"))


def generar_polizas_almacen(
    xlsx_path: str,
    output_dir: str | None = None,
    fecha_poliza: date | None = None,
    num_poliza_provision: int = 1,
    num_poliza_entrada: int = 1,
    num_poliza_salida: int = 1,
    tipo_poliza: int = TIPO_DIARIO,
    log_fn: Callable = print,
) -> dict:
    fecha_poliza = fecha_poliza or date.today()
    fecha_txt = fecha_poliza.strftime("%d/%m/%Y")
    out_dir = Path(output_dir) if output_dir else Path(xlsx_path).parent

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Datos Originales" not in wb.sheetnames:
        raise ValueError(
            "El archivo no tiene la hoja 'Datos Originales'. "
            "Usa el .xlsx generado por el botón COMPARAR."
        )
    ws = wb["Datos Originales"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header) if h}

    required = ["Sucursal", "Proveedor", "Factura", "Subtotal OC", "IVA OC", "Total OC"]
    faltantes = [c for c in required if c not in idx]
    if faltantes:
        raise ValueError(
            f"Faltan columnas {faltantes} en 'Datos Originales'. "
            "¿Ya corriste el RPA sobre este archivo antes de comparar?"
        )

    cuenta_cols = sorted(
        (i for h, i in idx.items() if h.startswith("Cuenta Contable")),
        key=lambda i: header[i],
    )

    almacen_catalog = load_almacen_catalog()

    mov_provision = ""
    mov_entrada = ""
    mov_salida = ""

    procesadas = 0
    con_poliza_sipp = 0
    sin_almacen: list[str] = []
    sucursales_sin_almacen: dict[str, None] = {}   # dict para preservar orden y deduplicar
    sin_cuenta_gasto: list[str] = []
    total_provision = 0.0
    total_entrada = 0.0
    total_salida = 0.0

    poliza_sipp_idx = idx.get("Poliza SIPP (JSON)")
    observaciones_idx = idx.get("Observaciones OC")
    folio_fiscal_idx = idx.get("Folio Fiscal")

    for row in ws.iter_rows(min_row=2, values_only=True):
        factura = str(row[idx["Factura"]] or "").strip()
        if not factura:
            continue

        sucursal = str(row[idx["Sucursal"]] or "").strip()
        proveedor = str(row[idx["Proveedor"]] or "").strip()
        subtotal_oc = _parse_amount(row[idx["Subtotal OC"]])
        iva_oc = _parse_amount(row[idx["IVA OC"]])
        folio_fiscal = str(row[folio_fiscal_idx] or "").strip() if folio_fiscal_idx is not None else ""

        if subtotal_oc <= 0:
            continue  # factura sin OC asociada — nada que contabilizar

        observaciones = str(row[observaciones_idx] or "").strip() if observaciones_idx is not None else ""
        concepto = observaciones if observaciones else f"{factura} - {proveedor}"

        # ── Provisión de Almacén ──────────────────────────────────────
        lineas_sipp = _parse_poliza_sipp(row[poliza_sipp_idx]) if poliza_sipp_idx is not None else []
        if lineas_sipp:
            for linea in lineas_sipp:
                if linea["cargo"] > 0:
                    mov_provision += _build_m_line(linea["cuenta"], 0, linea["cargo"], concepto, factura, folio_fiscal)
                    total_provision += linea["cargo"]
                if linea["abono"] > 0:
                    mov_provision += _build_m_line(linea["cuenta"], 1, linea["abono"], concepto, factura, folio_fiscal)
            con_poliza_sipp += 1
        else:
            mov_provision += _build_m_line(CUENTA_TRANSITO, 0, subtotal_oc, concepto, factura, folio_fiscal)
            abono_provision = subtotal_oc
            if iva_oc > 0:
                mov_provision += _build_m_line(CUENTA_IVA_DEFAULT, 0, iva_oc, concepto, factura, folio_fiscal)
                abono_provision += iva_oc
            mov_provision += _build_m_line(CUENTA_PROVEEDORES, 1, abono_provision, concepto, factura, folio_fiscal)
            total_provision += abono_provision

        # ── Entrada / Salida de Almacén (requieren cuenta de almacén) ──
        # El campo "Subtotal OC" viene del modal de la Orden de Compra y para
        # facturas de tipo DISTRIBUCIÓN/Corporativo puede reflejar el total
        # agregado de TODA la OC, no de esta factura — visto hasta 20x más
        # grande que el monto real. Cuando hay póliza SIPP real, usamos el
        # cargo de su primera línea (el gasto/tránsito real de la factura)
        # en vez de Subtotal OC para Entrada/Salida.
        monto_real = (
            lineas_sipp[0]["cargo"]
            if lineas_sipp and lineas_sipp[0]["cargo"] > 0
            else subtotal_oc
        )

        cuenta_almacen = _match_almacen(sucursal, almacen_catalog)
        if not cuenta_almacen:
            sin_almacen.append(f"{factura} ({proveedor}) — sucursal '{sucursal}' sin almacén")
            if sucursal:
                sucursales_sin_almacen[sucursal] = None
        else:
            mov_entrada += _build_m_line(cuenta_almacen, 0, monto_real, concepto, factura, folio_fiscal)
            mov_entrada += _build_m_line(CUENTA_TRANSITO, 1, monto_real, concepto, factura, folio_fiscal)
            total_entrada += monto_real

            cuentas = [
                str(row[i]).strip() for i in cuenta_cols if row[i] and str(row[i]).strip()
            ]
            if not cuentas:
                sin_cuenta_gasto.append(f"{factura} ({proveedor}) — sin cuenta contable SIPP")
            else:
                mov_salida += _build_m_line(cuentas[0], 0, monto_real, concepto, factura, folio_fiscal)
                mov_salida += _build_m_line(cuenta_almacen, 1, monto_real, concepto, factura, folio_fiscal)
                total_salida += monto_real

        procesadas += 1

    if procesadas == 0:
        raise ValueError("No se encontraron facturas con datos de OC para generar la póliza.")

    fecha_tag = fecha_poliza.strftime("%Y%m%d")
    out_provision = out_dir / f"Poliza_Provision_Almacen_{fecha_tag}.txt"
    out_entrada = out_dir / f"Poliza_Entrada_Almacen_{fecha_tag}.txt"
    out_salida = out_dir / f"Poliza_Salida_Almacen_{fecha_tag}.txt"

    _write_poliza(out_provision, fecha_poliza, tipo_poliza, num_poliza_provision,
                  f"PROVISION DE ALMACEN {fecha_txt}", mov_provision)
    _write_poliza(out_entrada, fecha_poliza, tipo_poliza, num_poliza_entrada,
                  f"ENTRADA DE ALMACEN {fecha_txt}", mov_entrada)
    _write_poliza(out_salida, fecha_poliza, tipo_poliza, num_poliza_salida,
                  f"SALIDA DE ALMACEN {fecha_txt}", mov_salida)

    log_fn(f"Pólizas generadas a partir de {procesadas} factura(s).", "info")
    log_fn(f"  Provisión con póliza real SIPP: {con_poliza_sipp} / {procesadas}", "info")
    log_fn(f"  Provisión de Almacén : {out_provision.name}  (${total_provision:,.2f})", "ok")
    log_fn(f"  Entrada de Almacén   : {out_entrada.name}  (${total_entrada:,.2f})", "ok")
    log_fn(f"  Salida de Almacén    : {out_salida.name}  (${total_salida:,.2f})", "ok")
    if sin_almacen:
        log_fn(f"  Sin cuenta de almacén (Entrada/Salida omitidas): {len(sin_almacen)}", "warn")
        for o in sin_almacen[:20]:
            log_fn(f"    - {o}", "warn")
    if sin_cuenta_gasto:
        log_fn(f"  Sin cuenta contable SIPP (Salida omitida): {len(sin_cuenta_gasto)}", "warn")
        for o in sin_cuenta_gasto[:20]:
            log_fn(f"    - {o}", "warn")

    return {
        "procesadas": procesadas,
        "con_poliza_sipp": con_poliza_sipp,
        "provision": {"path": str(out_provision), "total": round(total_provision, 2)},
        "entrada": {"path": str(out_entrada), "total": round(total_entrada, 2)},
        "salida": {"path": str(out_salida), "total": round(total_salida, 2)},
        "sin_almacen": sin_almacen,
        "sucursales_sin_almacen": list(sucursales_sin_almacen.keys()),
        "sin_cuenta_gasto": sin_cuenta_gasto,
    }


def generar_poliza_individual(
    xlsx_path: str,
    output_dir: str | None = None,
    fecha_poliza: date | None = None,
    num_poliza: int = 1,
    tipo_poliza: int = TIPO_DIARIO,
    log_fn: Callable = print,
) -> dict:
    """Genera UNA póliza combinada a partir de la póliza SIPP real de cada
    factura, repartiendo las líneas de gasto global (segmento '99') entre las
    estaciones de la hoja 'Distrib. Calculada' (usando el cargo real de SIPP,
    no 'Monto Distribuido'), y abonando a la cuenta del proveedor individual
    en vez de la cuenta genérica de Proveedores/Proveedores Tránsito.

    Facturas sin póliza SIPP real, sin cuenta de estación para alguna de sus
    partes distribuidas, o sin match de proveedor en el catálogo, se omiten
    por completo (para no generar una póliza desbalanceada) y se reportan.
    """
    fecha_poliza = fecha_poliza or date.today()
    fecha_txt = fecha_poliza.strftime("%d/%m/%Y")
    out_dir = Path(output_dir) if output_dir else Path(xlsx_path).parent

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    hojas_requeridas = ["Datos Originales", "Distrib. Calculada"]
    faltantes_hojas = [h for h in hojas_requeridas if h not in wb.sheetnames]
    if faltantes_hojas:
        raise ValueError(
            f"El archivo no tiene la(s) hoja(s) {faltantes_hojas}. "
            "Usa el .xlsx generado por el botón COMPARAR."
        )

    ws_orig = wb["Datos Originales"]
    header = [c.value for c in ws_orig[1]]
    idx = {h: i for i, h in enumerate(header) if h}
    required = ["Sucursal", "Proveedor", "Factura", "Subtotal OC"]
    faltantes_cols = [c for c in required if c not in idx]
    if faltantes_cols:
        raise ValueError(
            f"Faltan columnas {faltantes_cols} en 'Datos Originales'. "
            "¿Ya corriste el RPA sobre este archivo antes de comparar?"
        )
    poliza_sipp_idx = idx.get("Poliza SIPP (JSON)")
    if poliza_sipp_idx is None:
        raise ValueError(
            "El archivo no tiene la columna 'Poliza SIPP (JSON)' — el esquema "
            "Individual requiere pólizas SIPP reales capturadas por el RPA."
        )

    # ── Distrib. Calculada → {factura: [(estacion, pct), ...]} ──────────
    ws_dist = wb["Distrib. Calculada"]
    dist_header = [c.value for c in ws_dist[1]]
    dist_idx = {h: i for i, h in enumerate(dist_header) if h}
    distrib_por_factura: dict[str, list[tuple[str, float]]] = {}
    for row in ws_dist.iter_rows(min_row=2, values_only=True):
        factura = str(row[dist_idx["Factura"]] or "").strip()
        if not factura:
            continue  # fila vacía o "TOTAL DISTRIBUIDO"
        estacion = str(row[dist_idx["Estación Distribuida"]] or "").strip()
        if _normalize(estacion) == _normalize("(sin distribución)"):
            estacion = str(row[dist_idx["Sucursal"]] or "").strip()
        pct = _parse_amount(row[dist_idx["% Distribución"]])
        if not estacion or pct <= 0:
            continue
        distrib_por_factura.setdefault(factura, []).append((estacion, pct))

    estaciones_catalog = load_estaciones_catalog()
    proveedores_catalog = load_proveedores_catalog()

    observaciones_idx = idx.get("Observaciones OC")
    fecha_factura_idx = idx.get("Fecha Factura")
    folio_fiscal_idx = idx.get("Folio Fiscal")

    movimientos_por_fecha: dict[date, str] = {}
    procesadas = 0
    total_poliza = 0.0
    sin_poliza_sipp: list[str] = []
    sin_estacion: list[str] = []
    estaciones_sin_cuenta: dict[str, None] = {}
    sin_proveedor: list[str] = []
    proveedores_sin_cuenta: dict[str, None] = {}
    sin_fecha_factura: list[str] = []

    for row in ws_orig.iter_rows(min_row=2, values_only=True):
        factura = str(row[idx["Factura"]] or "").strip()
        if not factura:
            continue

        sucursal = str(row[idx["Sucursal"]] or "").strip()
        proveedor = str(row[idx["Proveedor"]] or "").strip()
        subtotal_oc = _parse_amount(row[idx["Subtotal OC"]])
        if subtotal_oc <= 0:
            continue

        lineas_sipp = _parse_poliza_sipp(row[poliza_sipp_idx])
        if not lineas_sipp:
            sin_poliza_sipp.append(f"{factura} ({proveedor})")
            continue

        observaciones = (
            str(row[observaciones_idx] or "").strip() if observaciones_idx is not None else ""
        )
        concepto = observaciones if observaciones else f"{factura} - {proveedor}"
        folio_fiscal = str(row[folio_fiscal_idx] or "").strip() if folio_fiscal_idx is not None else ""

        fecha_raw = row[fecha_factura_idx] if fecha_factura_idx is not None else None
        fecha_mov, uso_fallback = _parse_fecha_factura(fecha_raw, fecha_poliza)
        if uso_fallback:
            sin_fecha_factura.append(f"{factura} ({proveedor})")

        distrib = distrib_por_factura.get(factura) or [(sucursal, 100.0)]

        # ── Resolver código de estación de cada parte distribuida ──
        partes_resueltas: list[tuple[str, float]] = []
        estacion_faltante = False
        for estacion, pct in distrib:
            codigo = _match_catalog(estacion, estaciones_catalog)
            if not codigo:
                estacion_faltante = True
                estaciones_sin_cuenta[estacion] = None
            else:
                partes_resueltas.append((codigo, pct))
        if estacion_faltante:
            sin_estacion.append(f"{factura} ({proveedor})")
            continue

        # ── Resolver proveedor individual ──
        cuenta_proveedor = _match_catalog(proveedor, proveedores_catalog)
        if not cuenta_proveedor:
            sin_proveedor.append(f"{factura} ({proveedor})")
            proveedores_sin_cuenta[proveedor] = None
            continue

        # ── Construir movimientos de esta factura ──
        mov_factura = ""
        cargo_total = 0.0
        abono_proveedor_total = 0.0
        for linea in lineas_sipp:
            cuenta = linea["cuenta"]
            # La cuenta capturada de SIPP puede venir con o sin guiones (ej.
            # "502-99-03-0002" o "50299030002") — se normaliza al formato
            # limpio (sin separadores) y se detecta el segmento "99" por
            # posición fija: 3 dígitos de grupo + 2 de "estación" + resto.
            cuenta_limpia = _clean_account(cuenta)
            es_gasto_global = (
                len(cuenta_limpia) == 11
                and cuenta_limpia[:3] == "502"
                and cuenta_limpia[3:5] == "99"
            )

            if linea["cargo"] > 0:
                cargo_total += linea["cargo"]
                if es_gasto_global and partes_resueltas:
                    # Los % de los catálogos de distribución no siempre suman
                    # exactamente 100.00% al redondear a 2 decimales — repartir
                    # cada parte de forma independiente puede dejar la factura
                    # descuadrada por unos centavos (y con ella la póliza
                    # completa del día, que Contpaq rechaza si no balancea).
                    # La última parte absorbe el residuo de redondeo para que
                    # la suma cuadre exacto con el cargo real de SIPP.
                    montos = []
                    acumulado = 0.0
                    for _, pct in partes_resueltas[:-1]:
                        m = round(linea["cargo"] * pct / 100, 2)
                        montos.append(m)
                        acumulado += m
                    montos.append(round(linea["cargo"] - acumulado, 2))
                    for (codigo, _), monto in zip(partes_resueltas, montos):
                        if monto == 0:
                            continue
                        nueva_cuenta = cuenta_limpia[:3] + codigo + cuenta_limpia[5:]
                        mov_factura += _build_m_line(nueva_cuenta, 0, monto, concepto, factura, folio_fiscal)
                else:
                    mov_factura += _build_m_line(cuenta, 0, linea["cargo"], concepto, factura, folio_fiscal)

            if linea["abono"] > 0:
                if cuenta_limpia in (_clean_account(CUENTA_PROVEEDORES), _clean_account(CUENTA_TRANSITO)):
                    abono_proveedor_total += linea["abono"]
                else:
                    mov_factura += _build_m_line(cuenta, 1, linea["abono"], concepto, factura, folio_fiscal)

        if abono_proveedor_total > 0:
            mov_factura += _build_m_line(cuenta_proveedor, 1, abono_proveedor_total, concepto, factura, folio_fiscal)

        movimientos_por_fecha[fecha_mov] = movimientos_por_fecha.get(fecha_mov, "") + mov_factura
        total_poliza += cargo_total
        procesadas += 1

    if sin_poliza_sipp:
        log_fn(f"  Sin póliza SIPP real (omitidas): {len(sin_poliza_sipp)}", "warn")
        for o in sin_poliza_sipp[:20]:
            log_fn(f"    - {o}", "warn")
    if sin_estacion:
        log_fn(f"  Sin cuenta de estación (omitidas): {len(sin_estacion)}", "warn")
        for o in sin_estacion[:20]:
            log_fn(f"    - {o}", "warn")
    if sin_proveedor:
        log_fn(f"  Sin cuenta de proveedor (omitidas): {len(sin_proveedor)}", "warn")
        for o in sin_proveedor[:20]:
            log_fn(f"    - {o}", "warn")
    if sin_fecha_factura:
        log_fn(
            f"  Sin 'Fecha Factura' válida (se usó {fecha_txt} como fallback): "
            f"{len(sin_fecha_factura)}",
            "warn",
        )
        for o in sin_fecha_factura[:20]:
            log_fn(f"    - {o}", "warn")

    if procesadas == 0:
        raise ValueError(
            "No se encontraron facturas con póliza SIPP real para generar la "
            "póliza Individual."
        )

    # ── Un ARCHIVO por cada fecha distinta — Contpaq importa una sola póliza
    # (un bloque P + sus M) por archivo; concatenar varios bloques P en un
    # mismo .txt rompe la importación a partir del segundo bloque. Mismo
    # NumPol en todos los archivos. ──
    archivos: list[dict] = []
    num_actual = num_poliza
    for fecha_bloque in sorted(movimientos_por_fecha.keys()):
        concepto_bloque = f"POLIZA INDIVIDUAL {fecha_bloque.strftime('%d/%m/%Y')}"
        fecha_tag_bloque = fecha_bloque.strftime("%Y%m%d")
        out_path = out_dir / f"Poliza_Individual_{fecha_tag_bloque}.txt"
        _write_poliza(out_path, fecha_bloque, tipo_poliza, num_actual,
                      concepto_bloque, movimientos_por_fecha[fecha_bloque])
        num_actual += 1
        lineas_m = [l for l in movimientos_por_fecha[fecha_bloque].split("\r\n") if l.startswith("M  ")]
        total_cargo = sum(_parse_amount(l.split()[3]) for l in lineas_m if l.split()[2] == "0")
        total_abono = sum(_parse_amount(l.split()[3]) for l in lineas_m if l.split()[2] == "1")
        archivos.append({"path": str(out_path), "fecha": fecha_bloque.isoformat(), "total": round(total_cargo, 2)})
        log_fn(f"  {out_path.name}  (${total_cargo:,.2f})", "ok")
        if round(total_cargo - total_abono, 2) != 0:
            log_fn(
                f"    ⚠ NO CUADRA: cargo ${total_cargo:,.2f} vs abono ${total_abono:,.2f} "
                f"(diferencia ${total_cargo - total_abono:,.2f}) — Contpaq puede rechazar este archivo.",
                "error",
            )

    log_fn(
        f"Póliza Individual generada a partir de {procesadas} factura(s) "
        f"en {len(archivos)} archivo(s) (uno por fecha).",
        "info",
    )

    fecha_tag = fecha_poliza.strftime("%Y%m%d")
    zip_path = out_dir / f"Polizas_Individual_{fecha_tag}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for a in archivos:
            zf.write(a["path"], arcname=Path(a["path"]).name)
    log_fn(f"  Comprimido: {zip_path.name} ({len(archivos)} archivo(s))", "ok")

    return {
        "procesadas": procesadas,
        "archivos": archivos,
        "zip_path": str(zip_path),
        "total": round(total_poliza, 2),
        "sin_poliza_sipp": sin_poliza_sipp,
        "sin_estacion": sin_estacion,
        "estaciones_sin_cuenta": list(estaciones_sin_cuenta.keys()),
        "sin_proveedor": sin_proveedor,
        "proveedores_sin_cuenta": list(proveedores_sin_cuenta.keys()),
        "sin_fecha_factura": sin_fecha_factura,
    }


def generar_poliza_proveedores(
    xlsx_path: str,
    output_dir: str | None = None,
    fecha_poliza: date | None = None,
    num_poliza: int = 1,
    tipo_poliza: int = TIPO_DIARIO,
    log_fn: Callable = print,
) -> dict:
    """Genera UNA póliza por proveedor identificado:
      - Cargo : cuentas de gasto + IVA/retenciones (de la póliza SIPP real).
                Las cuentas 502-99 se distribuyen por estación igual que en
                la póliza Individual.
      - Abono : cuenta individual del proveedor (Cuentas_Proveedores.csv).

    Genera un TXT por proveedor, comprimidos en un ZIP.
    """
    fecha_poliza = fecha_poliza or date.today()
    fecha_txt = fecha_poliza.strftime("%d/%m/%Y")
    out_dir = Path(output_dir) if output_dir else Path(xlsx_path).parent

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    hojas_requeridas = ["Datos Originales", "Distrib. Calculada"]
    faltantes_hojas = [h for h in hojas_requeridas if h not in wb.sheetnames]
    if faltantes_hojas:
        raise ValueError(
            f"El archivo no tiene la(s) hoja(s) {faltantes_hojas}. "
            "Usa el .xlsx generado por el botón COMPARAR."
        )

    ws_orig = wb["Datos Originales"]
    header = [c.value for c in ws_orig[1]]
    idx = {h: i for i, h in enumerate(header) if h}
    required = ["Sucursal", "Proveedor", "Factura", "Subtotal OC"]
    faltantes_cols = [c for c in required if c not in idx]
    if faltantes_cols:
        raise ValueError(
            f"Faltan columnas {faltantes_cols} en 'Datos Originales'. "
            "¿Ya corriste el RPA sobre este archivo antes de comparar?"
        )
    poliza_sipp_idx = idx.get("Poliza SIPP (JSON)")
    if poliza_sipp_idx is None:
        raise ValueError(
            "El archivo no tiene la columna 'Poliza SIPP (JSON)' — el esquema "
            "Proveedores requiere pólizas SIPP reales capturadas por el RPA."
        )

    # ── Distrib. Calculada → {factura: [(estacion, pct), ...]} ──────────
    ws_dist = wb["Distrib. Calculada"]
    dist_header = [c.value for c in ws_dist[1]]
    dist_idx = {h: i for i, h in enumerate(dist_header) if h}
    distrib_por_factura: dict[str, list[tuple[str, float]]] = {}
    for row in ws_dist.iter_rows(min_row=2, values_only=True):
        factura = str(row[dist_idx["Factura"]] or "").strip()
        if not factura:
            continue
        estacion = str(row[dist_idx["Estación Distribuida"]] or "").strip()
        if _normalize(estacion) == _normalize("(sin distribución)"):
            estacion = str(row[dist_idx["Sucursal"]] or "").strip()
        pct = _parse_amount(row[dist_idx["% Distribución"]])
        if not estacion or pct <= 0:
            continue
        distrib_por_factura.setdefault(factura, []).append((estacion, pct))

    estaciones_catalog = load_estaciones_catalog()
    proveedores_catalog = load_proveedores_catalog()
    observaciones_idx = idx.get("Observaciones OC")
    folio_fiscal_idx = idx.get("Folio Fiscal")

    movimientos_por_proveedor: dict[str, str] = {}
    procesadas = 0
    total_poliza = 0.0
    sin_poliza_sipp: list[str] = []
    sin_estacion: list[str] = []
    estaciones_sin_cuenta: dict[str, None] = {}
    sin_proveedor: list[str] = []
    proveedores_sin_cuenta: dict[str, None] = {}

    for row in ws_orig.iter_rows(min_row=2, values_only=True):
        factura = str(row[idx["Factura"]] or "").strip()
        if not factura:
            continue

        sucursal = str(row[idx["Sucursal"]] or "").strip()
        proveedor = str(row[idx["Proveedor"]] or "").strip()
        subtotal_oc = _parse_amount(row[idx["Subtotal OC"]])
        if subtotal_oc <= 0:
            continue

        lineas_sipp = _parse_poliza_sipp(row[poliza_sipp_idx])
        if not lineas_sipp:
            sin_poliza_sipp.append(f"{factura} ({proveedor})")
            continue

        observaciones = (
            str(row[observaciones_idx] or "").strip() if observaciones_idx is not None else ""
        )
        concepto = observaciones if observaciones else f"{factura} - {proveedor}"
        folio_fiscal = str(row[folio_fiscal_idx] or "").strip() if folio_fiscal_idx is not None else ""

        distrib = distrib_por_factura.get(factura) or [(sucursal, 100.0)]

        # ── Resolver código de estación ──
        partes_resueltas: list[tuple[str, float]] = []
        estacion_faltante = False
        for estacion, pct in distrib:
            codigo = _match_catalog(estacion, estaciones_catalog)
            if not codigo:
                estacion_faltante = True
                estaciones_sin_cuenta[estacion] = None
            else:
                partes_resueltas.append((codigo, pct))
        if estacion_faltante:
            sin_estacion.append(f"{factura} ({proveedor})")
            continue

        # ── Resolver cuenta de proveedor individual ──
        cuenta_proveedor = _match_catalog(proveedor, proveedores_catalog)
        if not cuenta_proveedor:
            sin_proveedor.append(f"{factura} ({proveedor})")
            proveedores_sin_cuenta[proveedor] = None
            continue

        # ── Construir movimientos: cargo de SIPP, abono a proveedor ──
        mov_factura = ""
        cargo_total = 0.0
        abono_total = 0.0

        for linea in lineas_sipp:
            cuenta = linea["cuenta"]
            cuenta_limpia = _clean_account(cuenta)
            es_gasto_global = (
                len(cuenta_limpia) == 11
                and cuenta_limpia[:3] == "502"
                and cuenta_limpia[3:5] == "99"
            )

            if linea["cargo"] > 0:
                cargo_total += linea["cargo"]
                if es_gasto_global and partes_resueltas:
                    montos = []
                    acumulado = 0.0
                    for _, pct in partes_resueltas[:-1]:
                        m = round(linea["cargo"] * pct / 100, 2)
                        montos.append(m)
                        acumulado += m
                    montos.append(round(linea["cargo"] - acumulado, 2))
                    for (codigo, _), monto in zip(partes_resueltas, montos):
                        if monto == 0:
                            continue
                        nueva_cuenta = cuenta_limpia[:3] + codigo + cuenta_limpia[5:]
                        mov_factura += _build_m_line(nueva_cuenta, 0, monto, concepto, factura, folio_fiscal)
                else:
                    mov_factura += _build_m_line(cuenta, 0, linea["cargo"], concepto, factura, folio_fiscal)

            if linea["abono"] > 0:
                # Todos los abonos de SIPP se consolidan en la cuenta del proveedor
                abono_total += linea["abono"]

        if abono_total > 0:
            mov_factura += _build_m_line(cuenta_proveedor, 1, abono_total, concepto, factura, folio_fiscal)

        movimientos_por_proveedor[proveedor] = movimientos_por_proveedor.get(proveedor, "") + mov_factura
        total_poliza += cargo_total
        procesadas += 1

    if sin_poliza_sipp:
        log_fn(f"  Sin póliza SIPP real (omitidas): {len(sin_poliza_sipp)}", "warn")
        for o in sin_poliza_sipp[:20]:
            log_fn(f"    - {o}", "warn")
    if sin_estacion:
        log_fn(f"  Sin cuenta de estación (omitidas): {len(sin_estacion)}", "warn")
        for o in sin_estacion[:20]:
            log_fn(f"    - {o}", "warn")
    if sin_proveedor:
        log_fn(f"  Sin cuenta de proveedor (omitidas): {len(sin_proveedor)}", "warn")
        for o in sin_proveedor[:20]:
            log_fn(f"    - {o}", "warn")

    if procesadas == 0:
        raise ValueError(
            "No se encontraron facturas con póliza SIPP real para generar la "
            "póliza Proveedores."
        )

    archivos: list[dict] = []
    num_actual = num_poliza
    for proveedor_nombre in sorted(movimientos_por_proveedor.keys()):
        movimientos = movimientos_por_proveedor[proveedor_nombre]
        concepto_poliza = f"PROVEEDORES {_clean_concept(proveedor_nombre, 55)} {fecha_txt}"
        safe_nombre = re.sub(r'[<>:"/\\|?*]', "_", proveedor_nombre)[:50]
        fecha_tag = fecha_poliza.strftime("%Y%m%d")
        out_path = out_dir / f"Poliza_Proveedores_{safe_nombre}_{fecha_tag}.txt"
        _write_poliza(out_path, fecha_poliza, tipo_poliza, num_actual, concepto_poliza, movimientos)
        num_actual += 1

        lineas_m = [l for l in movimientos.split("\r\n") if l.startswith("M  ")]
        total_cargo = sum(_parse_amount(l.split()[3]) for l in lineas_m if len(l.split()) > 3 and l.split()[2] == "0")
        total_abono = sum(_parse_amount(l.split()[3]) for l in lineas_m if len(l.split()) > 3 and l.split()[2] == "1")
        archivos.append({"path": str(out_path), "proveedor": proveedor_nombre, "total": round(total_cargo, 2)})
        log_fn(f"  {out_path.name}  (${total_cargo:,.2f})", "ok")
        if round(total_cargo - total_abono, 2) != 0:
            log_fn(
                f"    ⚠ NO CUADRA: cargo ${total_cargo:,.2f} vs abono ${total_abono:,.2f} "
                f"(diferencia ${total_cargo - total_abono:,.2f}) — Contpaq puede rechazar este archivo.",
                "error",
            )

    fecha_tag = fecha_poliza.strftime("%Y%m%d")
    zip_path = out_dir / f"Polizas_Proveedores_{fecha_tag}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for a in archivos:
            zf.write(a["path"], arcname=Path(a["path"]).name)
    log_fn(f"  Comprimido: {zip_path.name} ({len(archivos)} proveedor(es))", "ok")

    log_fn(
        f"Póliza Proveedores generada a partir de {procesadas} factura(s) "
        f"en {len(archivos)} proveedor(es).",
        "info",
    )

    return {
        "procesadas": procesadas,
        "archivos": archivos,
        "zip_path": str(zip_path),
        "total": round(total_poliza, 2),
        "sin_poliza_sipp": sin_poliza_sipp,
        "sin_estacion": sin_estacion,
        "estaciones_sin_cuenta": list(estaciones_sin_cuenta.keys()),
        "sin_proveedor": sin_proveedor,
        "proveedores_sin_cuenta": list(proveedores_sin_cuenta.keys()),
    }

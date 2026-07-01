"""
compare_sucursales.py
Compara 'Observaciones OC' contra la lista de sucursales Petroplazas.

Estatus posibles:
  MATCH        — 1 sucursal detectada y coincide con Grupo Centro de Costo
  MISMATCH     — 1 sucursal detectada pero NO coincide con Grupo CC
  DISTRIBUCIÓN — 2+ sucursales detectadas (el total debe distribuirse entre ellas)
  SIN SUCURSAL — ninguna sucursal detectada en las observaciones

Uso:
    python3 compare_sucursales.py
    python3 compare_sucursales.py ruta/al/archivo.csv
"""

import csv
import re
import sys
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Catálogo de sucursales — nombres canónicos (tal como aparecen en SIPP) ──
# Estos son los nombres que se muestran en el reporte y se comparan con Grupo CC.
SUCURSALES = [
    "PATRIA", "BELLAVISTA", "PAPALOTE", "MADERO", "HEROICO", "PERICOS",
    "SAN PEDRO", "HUMAYA", "EL DIEZ", "MALECON", "OBREGON", "DOS VALLES",
    "FATIMA", "LOS ALAMOS", "REVOLUCION", "CUAUHTEMOC", "ALHUEY 1", "ALHUEY 2",
    "SABALO", "CERRITOS", "REAL DEL VALLE", "MARINA", "PLAYAS", "LAS HABAS",
    "FORESTA", "12 DE MAYO", "COLOSIO", "URBI", "FLORES MAGON", "SANTA ROSA",
    "GRIJALVA", "CARRASCO", "ROSARIO", "ESCUINAPA", "CONCORDIA",
    "HABAL", "CARDONES", "AEROPUERTO", "LA COLORADA", "LA PAZ",
    "MUNICH II", "EL DELFIN", "URBIVILLA", "EL CONCHI", "FORJADORES", "GOBERNADOR",
    "ESTADIO",
]

# ── Variantes textuales → nombre canónico ────────────────────────────────
# Cuando las observaciones usan una escritura alternativa, este dict la
# mapea al nombre canónico de SUCURSALES para que la detección y el
# reporte sean consistentes.
# Agregar nuevas variantes aquí sin tocar SUCURSALES.
ALIASES: dict = {
    # Bella Vista con espacio ↔ junto
    "BELLA VISTA":    "BELLAVISTA",
    # Habas sin artículo
    "HABAS":          "LAS HABAS",
    # München: número arábigo ↔ romano  (manejado también por _normalize_numerals)
    "MUNICH 2":       "MUNICH II",
    # München sin espacio (como aparece en Grupo CC: "ES_MunichII")
    "MUNICHII":       "MUNICH II",
    "MUNICH2":        "MUNICH II",
    # Grijalva con / sin sufijo UAS
    "GRIJALVA UAS":   "GRIJALVA",
    # Los Alamos sin artículo
    "ALAMOS":         "LOS ALAMOS",
    # Real del Valle / Real de Valle
    "REAL DE VALLE":  "REAL DEL VALLE",
    # Alhuey genérico (sin número) — si aparece solo, se asigna a ALHUEY 1
    "ALHUEY":         "ALHUEY 1",
    # La Colorada sin artículo
    "COLORADA":       "LA COLORADA",
    "DOS VALLE":    "DOS VALLES",
    "CUAUNTEMOC":    "CUAUHTEMOC",
    "DELFIN":       "EL DELFIN",


}

# ── Columnas del CSV (0-indexed desde la fila de encabezado) ────────────
COL_SUCURSAL        = 1
COL_FACTURA         = 3
COL_FOLIO           = 20
COL_GRUPO_CC        = 26
COL_TOTAL_MX        = 10   # Total en pesos MX (columna K original)
COL_CC_OC           = 31
COL_OBS_OC          = 32
COL_SUBTOTAL_OC     = 33   # Subtotal OC — columna AH (RPA), base para distribución
COL_CUENTA_CONTABLE = 38   # Cuenta Contable Factura — columna AM (RPA)
HEADER_ROW_IDX = 7   # fila 8 del CSV (0-indexed)

# ── Paleta de colores ────────────────────────────────────────────────────
FILL_GREEN      = PatternFill("solid", fgColor="C6EFCE")   # verde  — MATCH
FILL_RED        = PatternFill("solid", fgColor="FFC7CE")   # rojo   — MISMATCH / Sin cuenta
FILL_BLUE       = PatternFill("solid", fgColor="BDD7EE")   # azul   — DISTRIBUCIÓN
FILL_GRAY       = PatternFill("solid", fgColor="D9D9D9")   # gris   — SIN SUCURSAL
FILL_HEADER     = PatternFill("solid", fgColor="00264D")   # azul oscuro brand
FILL_CC_CATALOG  = PatternFill("solid", fgColor="E2EFDA")   # verde pálido  — Cuenta Catálogo
FILL_YELLOW_CC   = PatternFill("solid", fgColor="FFEB9C")   # amarillo      — Sin dato SIPP
FONT_YELLOW_CC   = Font(color="9C5700", bold=True, size=10) # naranja oscuro — Sin dato SIPP
FILL_ORANGE_CC   = PatternFill("solid", fgColor="FCE4D6")   # naranja pálido — Cuadre
FONT_ORANGE_CC   = Font(color="C55A11", bold=True, size=10) # naranja oscuro — Cuadre

FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
FONT_GREEN  = Font(color="276221", bold=True, size=10)
FONT_RED    = Font(color="9C0006", bold=True, size=10)
FONT_BLUE   = Font(color="1F4E79", bold=True, size=10)
FONT_NORMAL = Font(size=10)

THIN   = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Catálogo de cuentas contables ────────────────────────────────────────

CUENTAS_GASTOS_CSV      = Path(__file__).parent / "CuentasContables" / "Cuentas_Gastos.csv"
CUENTAS_PROVEEDORES_CSV = Path(__file__).parent / "CuentasContables" / "Cuentas_Proveedores.csv"


def _load_cuentas_csv(path: Path) -> dict:
    """Lee un CSV de cuentas (Cuenta, Nombre) y retorna {dígitos: (código, nombre)}."""
    result: dict = {}
    if not path.exists():
        return result
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            code = (row.get("Cuenta") or "").strip()
            name = (row.get("Nombre") or "").strip()
            if code:
                digits = re.sub(r"\D", "", code)
                if digits:
                    result[digits] = (code, name)
    return result


def load_cuentas_gastos() -> dict:
    return _load_cuentas_csv(CUENTAS_GASTOS_CSV)


def load_cuentas_proveedores() -> dict:
    return _load_cuentas_csv(CUENTAS_PROVEEDORES_CSV)


def load_all_cuentas() -> dict:
    """Combina gastos + proveedores en un solo catálogo de búsqueda."""
    catalog = _load_cuentas_csv(CUENTAS_GASTOS_CSV)
    catalog.update(_load_cuentas_csv(CUENTAS_PROVEEDORES_CSV))
    return catalog


def _lookup_cuenta(cuenta_factura: str, catalog: dict) -> tuple:
    """
    Busca cuenta_factura en el catálogo.
    Retorna (código_catálogo, nombre_catálogo, resultado) donde resultado es:
      'Match'         — SIPP devolvió cuenta Y existe en el catálogo
      'Mismatch'      — SIPP devolvió cuenta pero NO existe en el catálogo
      'Cuadre'        — cuenta de cuadre contable (_CUA-DRE-- u otro con letras/guión bajo)
      'Sin dato SIPP' — SIPP no devolvió ninguna cuenta
    """
    if not cuenta_factura or not cuenta_factura.strip():
        return ("", "", "Sin dato SIPP")
    raw = cuenta_factura.strip()
    # Detectar cuenta cuadre: "CUADRE" presente al quitar guiones/underscores/espacios
    # Ej: "_CUA-DRE--000000" → "CUADRE000000" → contiene "CUADRE"
    if "CUADRE" in re.sub(r"[-_\s]", "", raw).upper():
        return ("", "", "Cuadre")
    # Guarda contra strings de float como "50299040046.0" → "50299040046"
    raw = re.sub(r"\.0+$", "", raw)
    digits = re.sub(r"\D", "", raw)
    if digits and digits in catalog:
        code, name = catalog[digits]
        return (code, name, "Match")
    return ("", "", "Mismatch")


# ── Catálogos de distribución ─────────────────────────────────────────────

DISTRIBUCION_DIR = Path(__file__).parent / "Distribucion"

# Zona detectada (sin prefijo "ZONA ") → clave GCC en el catálogo
_ZONA_CATALOG_MAP: dict = {
    "CLN":      "CULIACAN",       # ZONA CLN      → CULIACAN
    "MAZATLAN": "MAZATLAN GRAL",  # ZONA MAZATLAN → Mazatlan_General (genérico)
    # Abreviaturas MZT
    "MZT":      "MAZATLAN GRAL",  # ZONA MZT      → Mazatlan_General
    "MZT 1":    "MAZATLAN 1",
    "MZT 2":    "MAZATLAN 2",
    "MZT 3":    "MAZATLAN 3",
    "MZT 4":    "MAZATLAN 4",
}


def _parse_amount(s: str) -> float:
    """'$1,393.00'  →  1393.0"""
    if not s:
        return 0.0
    try:
        return float(re.sub(r"[\$,\s]", "", s))
    except ValueError:
        return 0.0


def load_catalogs(directory: Path) -> dict:
    """
    Lee todos los CSV de la carpeta Distribucion/.
    Devuelve dict: clave_normalizada → [(estacion, float_pct), ...]
    Indexado tanto por GCC original normalizado como por GCC sin prefijo "ZONA ".
    """
    catalog: dict = {}
    if not directory.exists():
        return catalog
    for csv_file in sorted(directory.glob("*.csv")):
        try:
            with open(csv_file, encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    gcc      = (row.get("(GCC)", "") or "").strip()
                    estacion = (row.get("ESTACION", "") or "").strip()
                    try:
                        pct = float(row.get("PORCENTAJE", 0))
                    except (ValueError, TypeError):
                        pct = 0.0
                    if not gcc or not estacion:
                        continue
                    key = _normalize(gcc)
                    catalog.setdefault(key, []).append((estacion, pct))
                    # También indexar sin prefijo "ZONA " para búsquedas directas
                    stripped = re.sub(r"^ZONA\s+", "", key).strip()
                    if stripped != key:
                        catalog.setdefault(stripped, []).append((estacion, pct))
        except Exception:
            pass
    return catalog


def _zone_to_catalog_key(zone_label: str) -> str:
    """
    "ZONA MAZATLAN 2" → "MAZATLAN 2"
    "ZONA CLN"        → "CULIACAN"
    "ZONA MAZATLAN"   → "MAZATLAN GRAL"
    "ZONA CULIACAN"   → "CULIACAN"
    """
    norm     = _normalize(zone_label)
    stripped = re.sub(r"^ZONA\s+", "", norm).strip()
    return _ZONA_CATALOG_MAP.get(stripped, stripped)


def calculate_distribution(
    detected: str, grupo_cc: str, total_mx: float, catalog: dict
) -> list[tuple[str, float, float]]:
    """
    Retorna [(estacion, porcentaje, monto)] para un registro DISTRIBUCIÓN.
    Prioridad: zonas detectadas → Grupo CC en catálogo → split igual entre estaciones.
    """
    if total_mx <= 0:
        return []

    parts    = [p.strip() for p in detected.split("/") if p.strip()]
    zones    = [p for p in parts if _normalize(p).startswith("ZONA ")]
    stations = [p for p in parts if not _normalize(p).startswith("ZONA ")]
    result: list[tuple[str, float, float]] = []

    if zones:
        n_zones = len(zones)
        for zone in zones:
            key     = _zone_to_catalog_key(zone)
            entries = catalog.get(key, [])
            if entries:
                for estacion, pct in entries:
                    monto = round((total_mx / n_zones) * pct / 100, 2)
                    result.append((estacion, pct, monto))
            else:
                result.append((f"[{zone}] sin catálogo", 0.0, 0.0))

    elif stations:
        # Sin zona: intentar Grupo CC en catálogo; si no, split igual
        gcc_key = _normalize(grupo_cc)
        entries = catalog.get(gcc_key, [])
        if entries:
            for estacion, pct in entries:
                result.append((estacion, pct, round(total_mx * pct / 100, 2)))
        else:
            n         = len(stations)
            pct_each  = round(100 / n, 2)
            mnt_each  = round(total_mx / n, 2)
            for stn in stations:
                result.append((stn, pct_each, mnt_each))

    else:
        # Fallback: Grupo CC directo en catálogo
        for estacion, pct in catalog.get(_normalize(grupo_cc), []):
            result.append((estacion, pct, round(total_mx * pct / 100, 2)))

    return result


# ── Helpers de normalización ─────────────────────────────────────────────

def _normalize(text: str) -> str:
    """Quita acentos y convierte a mayúsculas."""
    if not text:
        return ""
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return text.upper().strip()


def _normalize_numerals(text: str) -> str:
    """Convierte numerales romanos comunes a arábigos para comparación.
    Ej: 'MUNICH II' → 'MUNICH 2', 'ALHUEY 1' no cambia.
    """
    for roman, arabic in [("VIII","8"),("VII","7"),("VI","6"),
                          ("IV","4"),("III","3"),("II","2")]:
        text = re.sub(r'\b' + roman + r'\b', arabic, text)
    return text


def _build_search_index() -> list:
    """
    Construye la lista de búsqueda combinando SUCURSALES + ALIASES.
    Cada entrada es (texto_normalizado, nombre_canónico).
    Ordenada de mayor a menor longitud para que los patrones más largos
    tengan prioridad ("LOS ALAMOS" antes que "ALAMOS").
    Las entradas de SUCURSALES tienen prioridad sobre ALIASES en caso
    de misma longitud.
    """
    seen: set = set()
    items: list = []

    for canonical in SUCURSALES:
        norm = _normalize(canonical)
        if norm not in seen:
            seen.add(norm)
            items.append((norm, canonical))

    for alias, canonical in ALIASES.items():
        norm = _normalize(alias)
        if norm not in seen:
            seen.add(norm)
            items.append((norm, canonical))

    return sorted(items, key=lambda x: len(x[0]), reverse=True)


_SUCURSALES_NORM = _build_search_index()


# Prefijos de dirección: si el match va precedido de uno de estos, es parte
# de una colonia/calle y NO debe tomarse como nombre de sucursal.
# Ej: "COL. RICARDO FLORES MAGON" → "FLORES MAGON" queda excluido.
_ADDRESS_CTX = re.compile(r'\b(?:COL|AV|CALLE|BLVD|BOULEVARD|CARR)\.?\s+\w+\s*$')


def find_all_sucursales_in_obs(obs: str) -> list[tuple[str, str]]:
    """
    Devuelve TODAS las sucursales (lista SUCURSALES) encontradas en el texto
    de observaciones con límite de palabra.
    Retorna lista de (nombre_original, norm_name) sin duplicados en nombre canónico,
    ignorando matches que formen parte de una dirección postal (COL., AV., etc.).
    """
    norm_obs = _normalize(obs)
    found: list[tuple[str, str]] = []
    consumed: set[int] = set()       # posiciones ya ocupadas por matches previos
    seen_canonical: set[str] = set() # evita contar la misma estación dos veces

    for norm_s, original in _SUCURSALES_NORM:
        pattern = r"(?<![A-Z0-9])" + re.escape(norm_s) + r"(?![A-Z0-9])"
        for m in re.finditer(pattern, norm_obs):
            positions = set(range(m.start(), m.end()))
            if positions & consumed:
                continue           # ya cubierto por un match más largo/anterior
            # Ignorar si el match forma parte de una dirección (COL., AV., etc.)
            prefix_ctx = norm_obs[max(0, m.start() - 30):m.start()]
            if _ADDRESS_CTX.search(prefix_ctx):
                continue           # falso positivo de colonia/calle — no consumir
            consumed |= positions
            if original not in seen_canonical:
                seen_canonical.add(original)
                found.append((original, norm_s))
            break

    return found


_ZONA_STOP = re.compile(r"\b(?:CON|DE|DEL|PARA|Y|A)\b")


_UT_PATTERN = re.compile(r'\b((?:AU|CA)-\d+)\b', re.IGNORECASE)


def load_utilitario_catalogs(directory: Path) -> dict:
    """
    Lee CSV de Distribucion/Utilitarios/.
    Devuelve {código_utilitario_upper → clave_distribución_original}
    Ej: {"AU-112": "MAZATLAN GRAL", "AU-065": "ES_Corporativo", "AU-109": "ZONA CULIACAN"}
    """
    ut_dir = directory / "Utilitarios"
    result: dict = {}
    if not ut_dir.is_dir():
        return result
    for csv_file in sorted(ut_dir.glob("*.csv")):
        try:
            with open(csv_file, encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    dist_key = (row.get("Distribución") or row.get("Distribucion") or "").strip()
                    ut_code  = (row.get("Utilitario") or "").strip()
                    if dist_key and ut_code:
                        result[ut_code.upper()] = dist_key
        except Exception:
            pass
    return result


def find_utilitario_in_text(text: str, ut_catalog: dict):
    """
    Busca el primer código de utilitario (AU-xxx / CA-xxx) en el texto que
    esté en el catálogo. Retorna (código, clave_dist) o None.
    """
    for m in _UT_PATTERN.finditer(text or ""):
        code = m.group(1).upper()
        if code in ut_catalog:
            return code, ut_catalog[code]
    return None


def find_zones_in_obs(obs: str) -> list[str]:
    """
    Devuelve una lista de etiquetas de zona encontradas en las observaciones.
    Ej: ['ZONA MAZATLAN 2', 'ZONA CULIACAN']
    """
    norm_obs = _normalize(obs)
    zones: list[str] = []
    seen: set[str] = set()

    # Captura "ZONA NOMBRE" o "ZONA NOMBRE N" donde N es un dígito.
    # No captura palabras adicionales como meses ("MAYO") o artículos.
    for m in re.finditer(r"\bZONA\s+([A-Z]+(?:\s+\d+)?)", norm_obs):
        raw = m.group(0).strip()
        # Recortar tokens funcionales que se colaron al final
        tokens = raw.split()
        clean_tokens = []
        for t in tokens:
            if _ZONA_STOP.fullmatch(t):
                break
            clean_tokens.append(t)
        label = " ".join(clean_tokens)
        if label not in seen:
            seen.add(label)
            zones.append(label)

    return zones


def sucursal_matches_grupo(norm_sucursal: str, grupo_cc: str) -> bool:
    """
    True si la sucursal normalizada coincide con el Grupo Centro de Costo.
    - Elimina prefijo ES_/CC_ y reemplaza underscores por espacios.
      Ej: 'ES_SANTA_ROSA' → 'SANTA ROSA', 'ES_El_Diez' → 'EL DIEZ'
    - Normaliza numerales romanos en ambos lados antes de comparar.
      Ej: 'MUNICH II' y 'MUNICH 2' → ambos quedan 'MUNICH 2' → MATCH
    - Usa similitud ≥ 0.90 para variaciones ortográficas menores.
      Ej: 'BELLAVISTA' vs 'BELLA VISTA' → ratio 0.95 → MATCH
    """
    # Separar camelCase antes de normalizar: "ES_MunichII" → "ES_Munich II"
    grupo_cc_split = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', grupo_cc)
    norm_grupo = _normalize(grupo_cc_split)
    norm_grupo_clean = re.sub(r"^(ES|CC)_+", "", norm_grupo)
    norm_grupo_clean = norm_grupo_clean.replace("_", " ").strip()

    # Aplicar conversión de numerales romanos a ambos lados
    suc_num   = _normalize_numerals(norm_sucursal)
    grupo_num = _normalize_numerals(norm_grupo_clean)

    if (suc_num == grupo_num
            or suc_num in grupo_num
            or grupo_num in suc_num):
        return True

    return SequenceMatcher(None, suc_num, grupo_num).ratio() >= 0.90


# ── Lectura de archivos (CSV o XLSX) ─────────────────────────────────────

def load_csv(filepath: str):
    with open(filepath, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    header = rows[HEADER_ROW_IDX]
    data   = rows[HEADER_ROW_IDX + 1:]
    return header, [r for r in data if any(c.strip() for c in r)]


def _cell_to_str(c) -> str:
    """Convierte un valor de celda openpyxl a str sin perder precisión.
    float 50299040046.0 → '50299040046'  (evita el '.0' que rompe búsquedas de dígitos)
    """
    if c is None:
        return ""
    if isinstance(c, float) and c.is_integer():
        return str(int(c))
    return str(c)


def load_xlsx(filepath: str):
    """Lee un XLSX y devuelve (header, data) en el mismo formato que load_csv."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_cell_to_str(c) for c in row])
    wb.close()
    if len(rows) <= HEADER_ROW_IDX:
        return [], []
    header = rows[HEADER_ROW_IDX]
    data   = rows[HEADER_ROW_IDX + 1:]
    return header, [r for r in data if any(c.strip() for c in r)]


def load_file(filepath: str):
    """Dispatcher: usa load_xlsx para .xlsx/.xls y load_csv para el resto."""
    if Path(filepath).suffix.lower() in (".xlsx", ".xls"):
        return load_xlsx(filepath)
    return load_csv(filepath)


def safe_get(row, idx: int) -> str:
    try:
        return (row[idx] or "").strip()
    except IndexError:
        return ""


# ── Helpers de escritura Excel ───────────────────────────────────────────

def _font_for_fill(fill):
    if fill == FILL_GREEN:
        return FONT_GREEN
    if fill == FILL_RED:
        return FONT_RED
    if fill == FILL_BLUE:
        return FONT_BLUE
    return FONT_NORMAL


def _hcell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = FILL_HEADER
    c.font   = FONT_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def _dcell(ws, row, col, value, fill=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font_for_fill(fill) if fill else FONT_NORMAL
    c.fill      = fill if fill else PatternFill()
    c.alignment = Alignment(vertical="center", wrap_text=wrap)
    c.border    = BORDER
    return c


# ── Hoja principal ───────────────────────────────────────────────────────

def build_main_sheet(ws, data, ut_catalog: dict | None = None):
    COLS = [
        (COL_SUCURSAL, "Sucursal"),
        (COL_FACTURA,  "Factura"),
        (COL_FOLIO,    "Folio"),
        (COL_GRUPO_CC, "Grupo Centro de Costo"),
        (COL_CC_OC,    "CC OC"),
        (COL_OBS_OC,   "Observaciones OC"),
    ]
    EXTRA = ["Sucursales Detectadas", "Resultado"]

    ws.title = "Comparación"
    ws.row_dimensions[1].height = 30

    for out_col, (_, name) in enumerate(COLS, start=1):
        _hcell(ws, 1, out_col, name)
    for i, name in enumerate(EXTRA, start=len(COLS) + 1):
        _hcell(ws, 1, i, name)

    widths = [18, 15, 15, 30, 22, 70, 30, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    counts  = {"MATCH": 0, "MISMATCH": 0, "DISTRIBUCIÓN": 0, "SIN SUCURSAL": 0}
    details = []

    for row_data in data:
        obs      = safe_get(row_data, COL_OBS_OC)
        cc_oc    = safe_get(row_data, COL_CC_OC)
        grupo_cc = safe_get(row_data, COL_GRUPO_CC)

        found = find_all_sucursales_in_obs(obs)   # [(nombre, norm), ...]
        zones = find_zones_in_obs(obs)             # ["ZONA MAZATLAN 2", ...]

        n_suc  = len(found)
        n_zone = len(zones)
        ut_dist_key = None   # clave de distribución por utilitario

        if n_suc == 0 and n_zone == 0:
            # Buscar utilitario en CC OC y en Observaciones OC
            ut_result = None
            if ut_catalog:
                ut_result = (find_utilitario_in_text(cc_oc, ut_catalog)
                             or find_utilitario_in_text(obs, ut_catalog))
            if ut_result:
                ut_code, ut_dist_key = ut_result
                detected = ut_code
                fill     = FILL_BLUE
                label    = "DISTRIBUCIÓN (UT)"
                counts["DISTRIBUCIÓN"] += 1
            elif _normalize(grupo_cc).startswith("ES_CORPORATIVO"):
                # Fallback: Petroplazas Corporativo sin sucursal detectada
                # → distribuir entre todas las estaciones de Corporativo.csv
                ut_dist_key = "ES_Corporativo"
                detected    = "ES_Corporativo"
                fill        = FILL_BLUE
                label       = "DISTRIBUCIÓN (Corp)"
                counts["DISTRIBUCIÓN"] += 1
            else:
                fill     = FILL_GRAY
                label    = "Sin sucursal"
                detected = ""
                counts["SIN SUCURSAL"] += 1

        elif n_suc == 1 and n_zone == 0:
            # Un solo match → MATCH o MISMATCH
            original, norm_s = found[0]
            detected = original
            if sucursal_matches_grupo(norm_s, grupo_cc):
                fill  = FILL_GREEN
                label = "MATCH ✓"
                counts["MATCH"] += 1
            else:
                fill  = FILL_RED
                label = "MISMATCH ✗"
                counts["MISMATCH"] += 1

        else:
            # Múltiples sucursales O al menos una zona → DISTRIBUCIÓN
            parts = [o for o, _ in found] + zones
            detected = " / ".join(parts)
            total    = n_suc + n_zone
            fill     = FILL_BLUE
            label    = f"DISTRIBUCIÓN ({total})"
            counts["DISTRIBUCIÓN"] += 1

        details.append({
            "sucursal":    safe_get(row_data, COL_SUCURSAL),
            "factura":     safe_get(row_data, COL_FACTURA),
            "folio":       safe_get(row_data, COL_FOLIO),
            "grupo_cc":    grupo_cc,
            "cc_oc":       cc_oc,
            "obs":         obs,
            "detected":    detected,
            "label":       label,
            "fill":        fill,
            "n_suc":       len(found),
            "total_mx":    safe_get(row_data, COL_TOTAL_MX),
            "subtotal_oc": safe_get(row_data, COL_SUBTOTAL_OC),
            "ut_dist_key": ut_dist_key,
        })

    for r_idx, d in enumerate(details, start=2):
        ws.row_dimensions[r_idx].height = 45

        _dcell(ws, r_idx, 1, d["sucursal"])
        _dcell(ws, r_idx, 2, d["factura"])
        _dcell(ws, r_idx, 3, d["folio"])

        # Grupo CC se colorea con el mismo estatus (salvo gris — sin sucursal)
        grupo_fill = d["fill"] if d["fill"] != FILL_GRAY else None
        _dcell(ws, r_idx, 4, d["grupo_cc"], fill=grupo_fill)

        _dcell(ws, r_idx, 5, d["cc_oc"])
        _dcell(ws, r_idx, 6, d["obs"], wrap=True)
        _dcell(ws, r_idx, 7, d["detected"], fill=d["fill"] if d["fill"] != FILL_GRAY else None, wrap=True)
        _dcell(ws, r_idx, 8, d["label"], fill=d["fill"])

    return counts, details


# ── Hoja de resumen ──────────────────────────────────────────────────────

def build_summary_sheet(ws, counts, total):
    ws.title = "Resumen"

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="RESUMEN DE COMPARACIÓN SUCURSALES")
    c.fill = FILL_HEADER
    c.font = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    c2 = ws.cell(row=2, column=1,
                 value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    c2.font      = Font(italic=True, size=10)
    c2.alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Resultado", "Cantidad", "Porcentaje", "Descripción"], start=1):
        _hcell(ws, 4, col, h)

    table_rows = [
        ("MATCH ✓",            counts["MATCH"],         FILL_GREEN,
         "1 sucursal detectada y coincide con Grupo CC"),
        ("MISMATCH ✗",         counts["MISMATCH"],      FILL_RED,
         "1 sucursal detectada pero NO coincide con Grupo CC"),
        ("DISTRIBUCIÓN",       counts["DISTRIBUCIÓN"],  FILL_BLUE,
         "2+ sucursales detectadas — el total debe distribuirse entre ellas"),
        ("Sin sucursal",       counts["SIN SUCURSAL"],  FILL_GRAY,
         "No se detectó ninguna sucursal en las observaciones"),
        ("TOTAL",              total,                   FILL_HEADER,
         "Total de registros procesados"),
    ]

    for r, (label, cnt, fill, desc) in enumerate(table_rows, start=5):
        pct = f"{cnt / total * 100:.1f}%" if total else "—"
        ws.row_dimensions[r].height = 22
        for col, val in enumerate([label, cnt, pct, desc], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.font = (FONT_HEADER if fill == FILL_HEADER else _font_for_fill(fill)) or FONT_NORMAL
            c.alignment = Alignment(
                horizontal="center" if col < 4 else "left", vertical="center"
            )
            c.border = BORDER

    for col, w in zip("ABCD", [22, 12, 12, 60]):
        ws.column_dimensions[col].width = w


# ── Hoja de resumen de cuentas contables ─────────────────────────────────

def build_resumen_cuentas_sheet(ws, counts: dict, total: int):
    ws.title = "Resumen CuentasC"

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="RESUMEN DE CONCILIACIÓN CUENTAS CONTABLES")
    c.fill      = FILL_HEADER
    c.font      = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    c2 = ws.cell(row=2, column=1,
                 value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    c2.font      = Font(italic=True, size=10)
    c2.alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Resultado", "Cantidad", "Porcentaje", "Descripción"], start=1):
        _hcell(ws, 4, col, h)

    n_match        = counts.get("match", 0)
    n_mismatch     = counts.get("mismatch", 0)
    n_cuadre       = counts.get("cuadre", 0)
    n_sin_dato     = counts.get("sin_dato_sipp", 0)

    table_rows = [
        (
            "Match ✓",
            n_match,
            FILL_GREEN,
            FONT_GREEN,
            "Cuenta contable de SIPP encontrada en el catálogo (gastos o proveedores)",
        ),
        (
            "Mismatch ✗",
            n_mismatch,
            FILL_RED,
            FONT_RED,
            "SIPP devolvió una cuenta contable pero no existe en el catálogo",
        ),
        (
            "Cuadre",
            n_cuadre,
            FILL_ORANGE_CC,
            FONT_ORANGE_CC,
            "Cuenta de cuadre contable (_CUA-DRE-- u otra con letras) — no se busca en catálogo",
        ),
        (
            "Sin dato SIPP",
            n_sin_dato,
            FILL_YELLOW_CC,
            FONT_YELLOW_CC,
            "SIPP no devolvió ninguna cuenta contable para esta factura",
        ),
        (
            "TOTAL",
            total,
            FILL_HEADER,
            FONT_HEADER,
            "Total de registros procesados",
        ),
    ]

    for r, (label, cnt, fill, font, desc) in enumerate(table_rows, start=5):
        pct = f"{cnt / total * 100:.1f}%" if total else "—"
        ws.row_dimensions[r].height = 22
        for col, val in enumerate([label, cnt, pct, desc], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.font = font
            c.alignment = Alignment(
                horizontal="center" if col < 4 else "left", vertical="center"
            )
            c.border = BORDER

    for col, w in zip("ABCD", [22, 12, 12, 70]):
        ws.column_dimensions[col].width = w


# ── Hoja por sucursal ────────────────────────────────────────────────────

def build_sucursal_detail_sheet(ws, details):
    ws.title = "Por Sucursal"

    stats: dict = defaultdict(lambda: {"MATCH": 0, "MISMATCH": 0, "DISTRIBUCIÓN": 0})

    for d in details:
        if not d["detected"]:
            continue
        # Para DISTRIBUCIÓN cada sucursal detectada recibe un conteo
        nombres = [s.strip() for s in d["detected"].split("/")]
        for nombre in nombres:
            if d["label"].startswith("MATCH"):
                stats[nombre]["MATCH"] += 1
            elif d["label"].startswith("MISMATCH"):
                stats[nombre]["MISMATCH"] += 1
            else:
                stats[nombre]["DISTRIBUCIÓN"] += 1

    for col, h in enumerate(["Sucursal", "MATCH ✓", "MISMATCH ✗", "DISTRIBUCIÓN", "Total"], start=1):
        _hcell(ws, 1, col, h)

    for col, w in zip("ABCDE", [25, 12, 14, 16, 10]):
        ws.column_dimensions[col].width = w

    for r, (suc, s) in enumerate(sorted(stats.items()), start=2):
        total = s["MATCH"] + s["MISMATCH"] + s["DISTRIBUCIÓN"]
        vals  = [suc, s["MATCH"], s["MISMATCH"], s["DISTRIBUCIÓN"], total]
        fills = [None, FILL_GREEN if s["MATCH"] else None,
                 FILL_RED if s["MISMATCH"] else None,
                 FILL_BLUE if s["DISTRIBUCIÓN"] else None, None]
        for col, (val, fill) in enumerate(zip(vals, fills), start=1):
            c = ws.cell(row=r, column=col, value=val)
            if fill:
                c.fill = fill
                c.font = _font_for_fill(fill)
            else:
                c.font = FONT_NORMAL
            c.border    = BORDER
            c.alignment = Alignment(
                horizontal="left" if col == 1 else "center", vertical="center"
            )


# ── Hoja exclusiva de DISTRIBUCIÓN ──────────────────────────────────────

def build_distribucion_sheet(ws, details):
    ws.title = "Distribución"

    COLS = ["Sucursal", "Factura", "Folio", "Grupo CC", "CC OC",
            "Observaciones OC", "Sucursales Detectadas", "# Estaciones"]

    for col, h in enumerate(COLS, start=1):
        _hcell(ws, 1, col, h)

    widths = [18, 15, 15, 28, 22, 70, 45, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dist_rows = [d for d in details if d["fill"] == FILL_BLUE]
    for r_idx, d in enumerate(dist_rows, start=2):
        ws.row_dimensions[r_idx].height = 50
        n = d["n_suc"]
        for col, val in enumerate(
            [d["sucursal"], d["factura"], d["folio"], d["grupo_cc"],
             d["cc_oc"], d["obs"], d["detected"], n], start=1
        ):
            wrap = col in (6, 7)
            _dcell(ws, r_idx, col, val,
                   fill=FILL_BLUE if col in (7, 8) else None, wrap=wrap)


# ── Hoja de distribución calculada ──────────────────────────────────────

def build_distribucion_calculada_sheet(ws, details, catalog):
    ws.title = "Distrib. Calculada"

    COLS = [
        "Sucursal", "Factura", "Folio", "Grupo CC",
        "Zona / Estaciones Detectadas", "Total Factura (MX)",
        "Estación Distribuida", "% Distribución", "Monto Distribuido",
    ]
    for col, h in enumerate(COLS, start=1):
        _hcell(ws, 1, col, h)

    widths = [18, 15, 15, 26, 42, 18, 26, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    FILL_AMOUNT = PatternFill("solid", fgColor="E2EFDA")   # verde pálido — monto

    r = 2
    grand_total = 0.0
    sin_distribucion = []

    for d in details:
        if d["fill"] != FILL_BLUE:
            sin_distribucion.append(d)
            continue

        subtotal_oc = _parse_amount(d.get("subtotal_oc", ""))
        ut_key = d.get("ut_dist_key")
        if ut_key:
            # Utilitario: buscar la clave de distribución directamente en el catálogo
            entries = catalog.get(_normalize(ut_key), [])
            dist = [(est, pct, round(subtotal_oc * pct / 100, 2)) for est, pct in entries]
        else:
            dist = calculate_distribution(d["detected"], d["grupo_cc"], subtotal_oc, catalog)

        if not dist or sum(monto for _, _, monto in dist) <= 0:
            ws.row_dimensions[r].height = 28
            for col, val in enumerate([
                d["sucursal"], d["factura"], d["folio"], d["grupo_cc"],
                d["detected"], subtotal_oc if subtotal_oc else "",
                "— sin catálogo —", "100.00%", subtotal_oc,
            ], start=1):
                _dcell(ws, r, col, val, fill=FILL_GRAY if col == 7 else None)
            grand_total += subtotal_oc
            r += 1
        else:
            for estacion, pct, monto in dist:
                ws.row_dimensions[r].height = 25
                fill_monto = FILL_AMOUNT if monto > 0 else None
                for col, val in enumerate([
                    d["sucursal"], d["factura"], d["folio"], d["grupo_cc"],
                    d["detected"], subtotal_oc if subtotal_oc else "",
                    estacion, f"{pct:.2f}%", monto,
                ], start=1):
                    _dcell(ws, r, col, val,
                           fill=FILL_BLUE if col == 7 else (fill_monto if col == 9 else None))
                grand_total += monto
                r += 1

    # Facturas sin distribución (MATCH / MISMATCH / Sin sucursal) — se listan
    # al final con su propio subtotal completo, para que el total general de
    # esta hoja reconcilie con la suma de "Subtotal OC" de Datos Originales.
    for d in sin_distribucion:
        subtotal_oc = _parse_amount(d.get("subtotal_oc", ""))
        if subtotal_oc <= 0:
            continue
        ws.row_dimensions[r].height = 25
        for col, val in enumerate([
            d["sucursal"], d["factura"], d["folio"], d["grupo_cc"],
            d["sucursal"], subtotal_oc,
            "(sin distribución)", "100.00%", subtotal_oc,
        ], start=1):
            _dcell(ws, r, col, val, fill=FILL_AMOUNT if col == 9 else None)
        grand_total += subtotal_oc
        r += 1

    # Fila de total general
    ws.row_dimensions[r].height = 28
    for col in range(1, 10):
        c = ws.cell(row=r, column=col)
        c.fill   = FILL_HEADER
        c.font   = FONT_HEADER
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r, column=7, value="TOTAL DISTRIBUIDO")
    ws.cell(row=r, column=9, value=round(grand_total, 2))


# ── Hoja de datos originales ─────────────────────────────────────────────

# Columnas añadidas por el RPA (0-indexed en las filas de datos):
# 31=CC OC  32=Obs OC  33=Subtotal  34=Descuento  35=IVA  36=G.Envío  37=Total OC
# 38+ = Cuenta Contable 1, 2, 3... (dinámicas)
_RPA_COL_RANGE = range(31, 39)   # columnas fijas del RPA (fondo amarillo): CC..Folio Fiscal
_RPA_FILL      = PatternFill("solid", fgColor="FFF2CC")   # amarillo pálido


def build_datos_originales_sheet(
    ws, header: list, data: list, cuentas_catalog: dict | None = None
) -> dict:
    """
    Construye la hoja 'Datos Originales'.
    Cada celda de Cuenta Contable se colorea verde (Match) o rojo (Mismatch).
    Retorna {'match': N, 'mismatch': M, 'sin_dato_sipp': P} contando por celda.
    """
    ws.title = "Datos Originales"
    ws.freeze_panes = "A2"

    # ── Encabezados de columnas fuente ──
    ws.row_dimensions[1].height = 24
    for col_idx, h in enumerate(header, start=1):
        c = ws.cell(row=1, column=col_idx, value=h or f"Col{col_idx}")
        c.fill      = FILL_HEADER
        c.font      = FONT_HEADER
        c.border    = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Detectar columnas de Cuenta Contable por nombre de header (0-indexed)
    cc_col_indices = [
        i for i, h in enumerate(header)
        if "Cuenta Contable" in (h or "")
    ]

    # ── Datos ──
    catalog = cuentas_catalog or {}
    counts  = {"match": 0, "mismatch": 0, "cuadre": 0, "sin_dato_sipp": 0}

    for r_idx, row_data in enumerate(data, start=2):
        # Escribir todas las columnas fuente
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            c.font      = FONT_NORMAL
            c.border    = BORDER
            c.alignment = Alignment(vertical="center")
            if col_idx - 1 in _RPA_COL_RANGE:
                c.fill = _RPA_FILL

        # Colorear columnas de Cuenta Contable según catálogo
        has_any_cuenta = False
        for cc_i in cc_col_indices:
            val = safe_get(row_data, cc_i)
            if not val or not val.strip():
                continue
            has_any_cuenta = True
            _, _, resultado = _lookup_cuenta(val, catalog)
            if resultado == "Match":
                fill, font = FILL_GREEN, FONT_GREEN
                counts["match"] += 1
            elif resultado == "Cuadre":
                fill, font = FILL_ORANGE_CC, FONT_ORANGE_CC
                counts["cuadre"] += 1
            else:
                fill, font = FILL_RED, FONT_RED
                counts["mismatch"] += 1
            c = ws.cell(row=r_idx, column=cc_i + 1)   # openpyxl es 1-indexed
            c.fill = fill
            c.font = font

        if not has_any_cuenta:
            counts["sin_dato_sipp"] += 1

    # ── Anchos de columna ──
    n_cols = max(len(header), max((len(r) for r in data), default=0))
    for i in range(1, n_cols + 1):
        letter = get_column_letter(i)
        if i - 1 in _RPA_COL_RANGE:
            ws.column_dimensions[letter].width = 28
        elif i - 1 in cc_col_indices:
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = 14

    return counts


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    default_csv = (
        Path(__file__).parent
        / "Recepcion_Facturas"
        / "RecepcionFacturas_Petroplazas-Unicos.csv"
    )
    csv_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_csv

    if not csv_path.exists():
        print(f"[ERROR] No se encontró: {csv_path}")
        sys.exit(1)

    print(f"Leyendo: {csv_path.name} …")
    header, data = load_file(str(csv_path))
    print(f"  {len(data)} registros cargados.")

    catalog        = load_catalogs(DISTRIBUCION_DIR)
    ut_catalog     = load_utilitario_catalogs(DISTRIBUCION_DIR)
    all_cuentas    = load_all_cuentas()

    if catalog:
        print(f"  {len(catalog)} claves de catálogo cargadas desde {DISTRIBUCION_DIR.name}/")
    else:
        print(f"  [AVISO] No se encontró la carpeta {DISTRIBUCION_DIR} — sin distribución calculada.")
    if ut_catalog:
        print(f"  {len(ut_catalog)} utilitarios cargados desde {DISTRIBUCION_DIR.name}/Utilitarios/")
    if all_cuentas:
        print(f"  {len(all_cuentas)} cuentas contables cargadas (gastos + proveedores)")
    else:
        print(f"  [AVISO] No se encontraron archivos de cuentas contables.")

    wb = Workbook()
    ws_orig      = wb.active          # 1. Datos Originales
    ws_sum       = wb.create_sheet()  # 2. Resumen
    ws_sum_cc    = wb.create_sheet()  # 3. Resumen CuentasC
    ws_suc       = wb.create_sheet()  # 4. Por Sucursal
    ws_main      = wb.create_sheet()  # 5. Comparación
    ws_dist      = wb.create_sheet()  # 6. Distribución
    ws_dist_calc = wb.create_sheet()  # 7. Distrib. Calculada

    counts, details    = build_main_sheet(ws_main, data, ut_catalog)
    cc_counts          = build_datos_originales_sheet(ws_orig, header, data, all_cuentas)
    build_summary_sheet(ws_sum, counts, len(data))
    build_resumen_cuentas_sheet(ws_sum_cc, cc_counts, len(data))
    build_sucursal_detail_sheet(ws_suc, details)
    build_distribucion_sheet(ws_dist, details)
    build_distribucion_calculada_sheet(ws_dist_calc, details, catalog)

    out_path = csv_path.parent / (csv_path.stem + "_comparacion.xlsx")
    wb.save(str(out_path))

    total = len(data)
    print()
    print("══════════════════════════════════════════")
    print("  RESUMEN SUCURSALES")
    print(f"  Total registros   : {total}")
    print(f"  MATCH ✓           : {counts['MATCH']}  ({counts['MATCH']/total*100:.1f}%)")
    print(f"  MISMATCH ✗        : {counts['MISMATCH']}  ({counts['MISMATCH']/total*100:.1f}%)")
    print(f"  DISTRIBUCIÓN      : {counts['DISTRIBUCIÓN']}  ({counts['DISTRIBUCIÓN']/total*100:.1f}%)")
    print(f"  Sin sucursal      : {counts['SIN SUCURSAL']}  ({counts['SIN SUCURSAL']/total*100:.1f}%)")
    print()
    print("  RESUMEN CUENTAS CONTABLES")
    print(f"  Match ✓           : {cc_counts['match']}  ({cc_counts['match']/total*100:.1f}%)")
    print(f"  Mismatch ✗        : {cc_counts['mismatch']}  ({cc_counts['mismatch']/total*100:.1f}%)")
    print(f"  Cuadre            : {cc_counts.get('cuadre', 0)}  ({cc_counts.get('cuadre', 0)/total*100:.1f}%)")
    print(f"  Sin dato SIPP     : {cc_counts['sin_dato_sipp']}  ({cc_counts['sin_dato_sipp']/total*100:.1f}%)")
    print("══════════════════════════════════════════")
    print(f"  Archivo generado  : {out_path.name}")
    print("══════════════════════════════════════════")


if __name__ == "__main__":
    main()

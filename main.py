"""
RPA — Distribución Petroplazas | Recepción de Facturas
GUI built with Flet. Playwright automation runs as an async task on Flet's
own event loop (page.run_task) — no manual threading/queue needed.
"""

import argparse
import asyncio
import subprocess
import sys
import threading
from datetime import datetime
from pathlib import Path

import flet as ft

from rpa.catalog_io import read_catalog, write_catalog

# ── CLI args ─────────────────────────────────────────────
_parser = argparse.ArgumentParser(add_help=False)
_parser.add_argument("--max", type=int, default=None, metavar="N",
                     help="Limitar a los primeros N folios del Excel")
_args, _remaining = _parser.parse_known_args()
sys.argv = [sys.argv[0]] + _remaining
MAX_RECORDS: int | None = _args.max

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

ROOT_DIR = Path(__file__).parent
RECEPCION_DIR = ROOT_DIR / "Recepcion_Facturas"
DISTRIBUCION_DIR = ROOT_DIR / "Distribucion"
CUENTAS_DIR = ROOT_DIR / "CuentasContables"
LOGO_ICON_PATH = ROOT_DIR / "Logo_Petroil.png"      # ícono cuadrado (taskbar)
LOGO_WIDE_PATH = ROOT_DIR / "grupopetroil.png"      # logotipo horizontal (header)

# Acentos de marca (fijos, no cambian con el tema)
C_GOLD       = "#f0a500"
C_BTN_RUN    = "#1a5e35"
C_BTN_STOP   = "#7b0000"
C_BTN_CMP    = "#1a4a6e"
C_BTN_CAT    = "#3b2a6b"
C_BTN_POL    = "#6e4a1a"
C_STAT_OK    = "#1e7e34"
C_STAT_ERR   = "#c0392b"
C_STAT_PEN   = "#b8860b"

LOG_COLOR_DARK  = {"ok": "#4caf50", "error": "#ef5350", "warn": "#ffb300", "info": "#b0bec5"}
LOG_COLOR_LIGHT = {"ok": "#1e7e34", "error": "#c0392b", "warn": "#9a6b00", "info": "#4a5568"}
LOG_ICON  = {"ok": "✓", "error": "✗", "warn": "⚠", "info": "·"}

DIST_COLUMNS = ["(GCC)", "ZONA(CC)", "ESTACION", "PORCENTAJE"]
CATALOG_TABS = [
    {"name": "Mazatlán General",    "path": DISTRIBUCION_DIR / "Mazatlan_General.csv", "columns": DIST_COLUMNS},
    {"name": "Corporativo",         "path": DISTRIBUCION_DIR / "Corporativo.csv",       "columns": DIST_COLUMNS},
    {"name": "Zonas",               "path": DISTRIBUCION_DIR / "Zonas.csv",             "columns": DIST_COLUMNS},
    {"name": "Cuentas Gastos",      "path": CUENTAS_DIR / "Cuentas_Gastos.csv",         "columns": ["Cuenta", "Nombre"]},
    {"name": "Cuentas Proveedores", "path": CUENTAS_DIR / "Cuentas_Proveedores.csv",    "columns": ["Cuenta", "Nombre"]},
    {"name": "Almacén Ref",         "path": CUENTAS_DIR / "Cuentas_AlmacénRef.csv",     "columns": ["Cuenta", "Almacén"]},
    {"name": "Cuentas Gasto Estaciones", "path": CUENTAS_DIR / "Cuentas_GastoEstaciones.csv", "columns": ["Cuenta", "Nombre"]},
]


def _palette(is_dark: bool) -> dict:
    if is_dark:
        return dict(
            bg="#1e1e1e", surface="#2b2b2b", header_bg="#00264d",
            text="white", subtext="#99bbcc", border="#3a3a3a",
            log_bg="#161616", input_fill="#1e1e1e",
        )
    return dict(
        bg="#eef1f6", surface="white", header_bg="white",
        text="#0a2f5c", subtext="#5a6b7c", border="#e1e6ee",
        log_bg="#f7f8fa", input_fill="#f7f9fb",
    )


# ════════════════════════════════════════════════════════
# Catálogos — vista de edición
# ════════════════════════════════════════════════════════
class CatalogTab:
    """Una pestaña: tabla editable respaldada por un catálogo CSV."""

    def __init__(self, page: ft.Page, spec: dict, colors: dict):
        self.page = page
        self.path: Path = spec["path"]
        self.columns: list[str] = spec["columns"]
        self.colors = colors
        self.rows: list[dict] = read_catalog(self.path, self.columns)
        self.selected_index: int | None = None

        self.data_table = ft.DataTable(
            columns=[ft.DataColumn(ft.Text(c, weight=ft.FontWeight.BOLD, color=colors["text"])) for c in self.columns],
            rows=[],
            border=ft.border.all(1, colors["border"]),
            heading_row_color=colors["bg"],
            heading_row_height=38,
            data_row_min_height=34,
        )
        self._rebuild_rows()

    def _rebuild_rows(self):
        data_rows = []
        for i, row in enumerate(self.rows):
            is_sel = i == self.selected_index
            data_rows.append(
                ft.DataRow(
                    cells=[ft.DataCell(ft.Text(row.get(c, ""), color=self.colors["text"])) for c in self.columns],
                    selected=is_sel,
                    on_select_changed=lambda e, idx=i: self._on_select(idx),
                    color="#1a5e8a" if is_sel else None,
                )
            )
        self.data_table.rows = data_rows

    def _on_select(self, idx: int):
        self.selected_index = idx
        self._rebuild_rows()
        self.page.update()

    def add_row(self):
        return self._row_dialog(is_new=True)

    def edit_row(self):
        if self.selected_index is None:
            self.page.open(ft.SnackBar(ft.Text("Selecciona una fila primero."), bgcolor="#7d6608"))
            return
        return self._row_dialog(is_new=False)

    def delete_row(self):
        if self.selected_index is None:
            self.page.open(ft.SnackBar(ft.Text("Selecciona una fila primero."), bgcolor="#7d6608"))
            return

        def _confirm(e):
            self.rows.pop(self.selected_index)
            self.selected_index = None
            self._rebuild_rows()
            dlg.open = False
            self.page.update()

        def _cancel(e):
            dlg.open = False
            self.page.update()

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("Eliminar fila"),
            content=ft.Text("¿Eliminar la fila seleccionada?"),
            actions=[
                ft.TextButton("Cancelar", on_click=_cancel),
                ft.TextButton("Eliminar", on_click=_confirm, style=ft.ButtonStyle(color="#ef5350")),
            ],
        )
        self.page.open(dlg)

    def _row_dialog(self, is_new: bool):
        current = {c: "" for c in self.columns} if is_new else dict(self.rows[self.selected_index])
        fields = {
            c: ft.TextField(label=c, value=current.get(c, ""), autofocus=(i == 0))
            for i, c in enumerate(self.columns)
        }

        def _ok(e):
            new_row = {c: fields[c].value.strip() for c in self.columns}
            if not new_row.get(self.columns[0]):
                fields[self.columns[0]].error_text = "Este campo es obligatorio"
                self.page.update()
                return
            if is_new:
                self.rows.append(new_row)
                self.selected_index = len(self.rows) - 1
            else:
                self.rows[self.selected_index] = new_row
            self._rebuild_rows()
            dlg.open = False
            self.page.update()

        def _cancel(e):
            dlg.open = False
            self.page.update()

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("Nueva fila" if is_new else "Editar fila"),
            content=ft.Column(list(fields.values()), tight=True, width=360),
            actions=[
                ft.TextButton("Cancelar", on_click=_cancel),
                ft.FilledButton("Aceptar", on_click=_ok),
            ],
        )
        self.page.open(dlg)

    def save(self):
        write_catalog(self.path, self.columns, self.rows)


def build_catalogs_view(
    page: ft.Page, colors: dict,
    initial_tab_index: int = 0,
    pending_almacen_rows: list[str] | None = None,
    pending_proveedor_rows: list[str] | None = None,
) -> ft.View:
    tabs_data = [CatalogTab(page, spec, colors) for spec in CATALOG_TABS]

    def _prefill(tab_name: str, name_col: str, values: list[str], label_fn=lambda v: v):
        tab = next((t for t, spec in zip(tabs_data, CATALOG_TABS) if spec["name"] == tab_name), None)
        if tab is None:
            return
        existing = {row.get(name_col, "").strip().upper() for row in tab.rows}
        for value in values:
            label = label_fn(value.strip())
            if label.strip().upper() not in existing:
                tab.rows.append({"Cuenta": "", name_col: label})
                existing.add(label.strip().upper())
        tab._rebuild_rows()

    if pending_almacen_rows:
        _prefill("Almacén Ref", "Almacén", pending_almacen_rows, lambda v: f"Sucursal ({v})")
    if pending_proveedor_rows:
        _prefill("Cuentas Proveedores", "Nombre", pending_proveedor_rows)

    def _save_all(e):
        try:
            for tab in tabs_data:
                tab.save()
            page.open(ft.SnackBar(ft.Text("Catálogos guardados correctamente."), bgcolor="#1a5e35"))
        except Exception as exc:
            page.open(ft.SnackBar(ft.Text(f"Error al guardar: {exc}"), bgcolor="#7b241c"))

    def _back(e):
        page.go("/")

    tabs = ft.Tabs(
        selected_index=initial_tab_index,
        tabs=[
            ft.Tab(
                text=spec["name"],
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Container(
                                ft.Column([tab.data_table], scroll=ft.ScrollMode.AUTO, expand=True),
                                expand=True, padding=8,
                            ),
                            ft.Row(
                                [
                                    ft.ElevatedButton("Agregar fila", icon=ft.Icons.ADD,
                                                      on_click=lambda e, t=tab: t.add_row()),
                                    ft.OutlinedButton("Editar", icon=ft.Icons.EDIT,
                                                      on_click=lambda e, t=tab: t.edit_row()),
                                    ft.OutlinedButton(
                                        "Eliminar", icon=ft.Icons.DELETE,
                                        on_click=lambda e, t=tab: t.delete_row(),
                                        style=ft.ButtonStyle(color="#ef5350"),
                                    ),
                                ],
                                spacing=8,
                            ),
                        ],
                        expand=True,
                    ),
                    padding=10,
                ),
            )
            for tab, spec in zip(tabs_data, CATALOG_TABS)
        ],
        expand=True,
    )

    header = ft.Container(
        ft.Row(
            [
                ft.Text("Catálogos de Distribución y Cuentas Contables",
                        size=16, weight=ft.FontWeight.BOLD, color=colors["text"]),
                ft.Container(expand=True),
                ft.Text("Selecciona una fila para editar",
                        size=11, color=colors["subtext"]),
            ],
        ),
        bgcolor=colors["header_bg"], padding=12,
        border=ft.border.only(bottom=ft.BorderSide(1, colors["border"])),
    )

    footer = ft.Row(
        [
            ft.OutlinedButton("Cerrar", icon=ft.Icons.CLOSE, on_click=_back),
            ft.FilledButton("Guardar todos", icon=ft.Icons.SAVE, on_click=_save_all,
                            style=ft.ButtonStyle(bgcolor=C_BTN_RUN, color="white")),
        ],
        alignment=ft.MainAxisAlignment.END,
        spacing=8,
    )

    return ft.View(
        "/catalogos",
        [header, ft.Container(height=3, bgcolor=C_GOLD), ft.Container(tabs, expand=True, padding=10), ft.Container(footer, padding=10)],
        padding=0,
        bgcolor=colors["bg"],
    )


# ════════════════════════════════════════════════════════
# App principal
# ════════════════════════════════════════════════════════
class RPAApp:
    def __init__(self, page: ft.Page):
        self.page = page
        page.title = "RPA — Recepción de Facturas | SIPP Petroplazas"
        page.window.width = 1080
        page.window.height = 820
        page.window.min_width = 920
        page.window.min_height = 700
        page.padding = 0
        if LOGO_ICON_PATH.exists():
            page.window.icon = str(LOGO_ICON_PATH)

        # ── Estado de la app (sobrevive a los rebuilds de la vista, p.ej.
        # al cambiar de tema o navegar a Catálogos y volver) ──
        self.is_running = False
        self.start_ts: datetime | None = None
        self.is_dark = False
        self.file_path = ""
        self.total = 0
        self.processed = 0
        self.errors = 0
        self.status_text = "En espera — selecciona un archivo."
        self.elapsed_text = ""
        self.progress_val = 0.0
        self.log_entries: list[tuple[str, str]] = []   # (texto_formateado, nivel)
        self.update_banner_visible = False
        self.update_banner_text = ""
        self.username = "afuentes"
        self.password = ""
        self.headless = False
        self.workers_selected = "1"
        self.auto_compare = False
        self.catalogs_initial_tab = 0
        self.pending_almacen_rows: list[str] | None = None
        self.pending_proveedor_rows: list[str] | None = None

        self.file_picker = ft.FilePicker(on_result=self._on_file_picked)
        self.compare_picker = ft.FilePicker(on_result=self._on_compare_picked)
        self.poliza_picker = ft.FilePicker(on_result=self._on_poliza_picked)
        page.overlay.extend([self.file_picker, self.compare_picker, self.poliza_picker])

        page.on_route_change = self._on_route_change
        page.go(page.route or "/")

        threading.Thread(target=self._check_updates_sync, daemon=True).start()

    @property
    def colors(self) -> dict:
        return _palette(self.is_dark)

    # ────────────────────────────────────────────────────
    # Routing
    # ────────────────────────────────────────────────────
    def _on_route_change(self, e: ft.RouteChangeEvent):
        self.page.theme_mode = ft.ThemeMode.DARK if self.is_dark else ft.ThemeMode.LIGHT
        self.page.bgcolor = self.colors["bg"]
        self.page.views.clear()
        if self.page.route == "/catalogos":
            self.page.views.append(build_catalogs_view(
                self.page, self.colors,
                initial_tab_index=self.catalogs_initial_tab,
                pending_almacen_rows=self.pending_almacen_rows,
                pending_proveedor_rows=self.pending_proveedor_rows,
            ))
            self.catalogs_initial_tab = 0
            self.pending_almacen_rows = None
            self.pending_proveedor_rows = None
        else:
            self.page.views.append(self._build_main_view())
        self.page.update()

    def _toggle_theme(self, e):
        self.is_dark = not self.is_dark
        self.page.go(self.page.route)

    # ────────────────────────────────────────────────────
    # Vista principal
    # ────────────────────────────────────────────────────
    def _build_main_view(self) -> ft.View:
        colors = self.colors
        self._build_header(colors)
        self._build_update_banner(colors)
        self._build_config(colors)
        self._build_stats(colors)
        self._build_progress(colors)
        self._build_logs(colors)
        self._build_buttons(colors)

        return ft.View(
            "/",
            [
                self.header,
                ft.Container(height=3, bgcolor=C_GOLD),
                self.update_banner,
                self.cfg_card,
                self.stats_row,
                self.progress_col,
                ft.Container(self.log_card, expand=True, padding=ft.padding.symmetric(horizontal=12)),
                ft.Container(self.buttons_row, padding=12),
            ],
            padding=0,
            spacing=10,
            scroll=ft.ScrollMode.AUTO,
            bgcolor=colors["bg"],
        )

    def _build_header(self, colors: dict):
        if LOGO_WIDE_PATH.exists():
            logo = ft.Image(src=str(LOGO_WIDE_PATH), height=38, fit=ft.ImageFit.CONTAIN)
        else:
            logo = ft.Text("⛽", size=22)

        self.btn_check_update = ft.IconButton(
            icon=ft.Icons.REFRESH, icon_color=colors["text"],
            tooltip="Comprobar actualizaciones",
            on_click=self._manual_check_updates,
        )
        self.btn_theme = ft.IconButton(
            icon=ft.Icons.DARK_MODE if not self.is_dark else ft.Icons.LIGHT_MODE,
            icon_color=colors["text"],
            tooltip="Cambiar tema",
            on_click=self._toggle_theme,
        )
        self.header = ft.Container(
            ft.Row(
                [
                    logo,
                    ft.Column(
                        [
                            ft.Text("Recepción de Facturas  |  SIPP Petroplazas",
                                    size=16, weight=ft.FontWeight.BOLD, color=colors["text"]),
                            ft.Text("Automatización de captura CC + Observaciones OC",
                                    size=11, color=colors["subtext"]),
                        ],
                        spacing=0, expand=True,
                    ),
                    self.btn_theme,
                    self.btn_check_update,
                ],
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            bgcolor=colors["header_bg"], padding=ft.padding.symmetric(horizontal=18, vertical=10),
            border=ft.border.only(bottom=ft.BorderSide(1, colors["border"])) if not self.is_dark else None,
        )

    def _build_update_banner(self, colors: dict):
        self.update_lbl = ft.Text(self.update_banner_text, color="white",
                                   weight=ft.FontWeight.BOLD, size=12, expand=True)
        self.btn_update = ft.ElevatedButton(
            "Actualizar y reiniciar", icon=ft.Icons.DOWNLOAD, on_click=self._do_update,
            style=ft.ButtonStyle(bgcolor="white", color="#1b5e33"),
        )
        self.update_banner = ft.Container(
            ft.Row(
                [self.update_lbl, self.btn_update,
                 ft.IconButton(icon=ft.Icons.CLOSE, icon_color="#aaddaa",
                               on_click=self._dismiss_update_banner)],
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            bgcolor="#1b5e33", padding=ft.padding.symmetric(horizontal=14, vertical=6),
            visible=self.update_banner_visible,
        )

    def _build_config(self, colors: dict):
        self.file_field = ft.TextField(
            label="Archivo Excel", read_only=True, expand=True, value=self.file_path,
            hint_text="Seleccionar archivo .xlsx ...",
            filled=True, fill_color=colors["input_fill"], border_color=colors["border"],
            color=colors["text"], label_style=ft.TextStyle(color=colors["subtext"]),
        )
        browse_btn = ft.ElevatedButton("Examinar", on_click=self._browse_file)

        self.user_field = ft.TextField(
            label="Usuario", value=self.username, width=180,
            filled=True, fill_color=colors["input_fill"], border_color=colors["border"],
            color=colors["text"], label_style=ft.TextStyle(color=colors["subtext"]),
            on_change=lambda e: setattr(self, "username", e.control.value),
        )
        self.pass_field = ft.TextField(
            label="Contraseña", password=True, can_reveal_password=True, width=180,
            value=self.password,
            filled=True, fill_color=colors["input_fill"], border_color=colors["border"],
            color=colors["text"], label_style=ft.TextStyle(color=colors["subtext"]),
            on_change=lambda e: setattr(self, "password", e.control.value),
        )
        self.headless_chk = ft.Checkbox(
            label="Modo silencioso (headless)", value=self.headless,
            label_style=ft.TextStyle(color=colors["text"]),
            on_change=lambda e: setattr(self, "headless", e.control.value),
        )

        self.workers_seg = ft.SegmentedButton(
            segments=[
                ft.Segment(value="1", label=ft.Text("1×")),
                ft.Segment(value="2", label=ft.Text("2×")),
                ft.Segment(value="4", label=ft.Text("4×")),
                ft.Segment(value="8", label=ft.Text("8×")),
            ],
            selected={self.workers_selected},
            allow_multiple_selection=False,
            allow_empty_selection=False,
            on_change=lambda e: setattr(self, "workers_selected", next(iter(self.workers_seg.selected))),
        )

        self.cfg_card = ft.Container(
            ft.Column(
                [
                    ft.Row([self.file_field, browse_btn], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    ft.Row(
                        [self.user_field, self.pass_field, self.headless_chk],
                        spacing=20, vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                    ft.Row(
                        [
                            ft.Text("Sesiones paralelas:", weight=ft.FontWeight.BOLD, color=colors["text"]),
                            self.workers_seg,
                            ft.Text(
                                "  ·  más sesiones = más velocidad (cada una inicia login en SIPP)",
                                size=10, color=colors["subtext"],
                            ),
                        ],
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                ],
                spacing=12,
            ),
            bgcolor=colors["surface"], border_radius=10, padding=16,
            margin=ft.margin.symmetric(horizontal=12),
            border=ft.border.all(1, colors["border"]),
            shadow=ft.BoxShadow(blur_radius=6, color="#14000000", offset=ft.Offset(0, 2)) if not self.is_dark else None,
        )

    def _stat_card(self, colors: dict, label: str, value: str, color: str) -> ft.Container:
        val_lbl = ft.Text(value, size=26, weight=ft.FontWeight.BOLD, color=color,
                          text_align=ft.TextAlign.CENTER)
        card = ft.Container(
            ft.Column(
                [
                    ft.Container(height=3, bgcolor=color,
                                border_radius=ft.border_radius.only(top_left=8, top_right=8)),
                    ft.Container(
                        ft.Column(
                            [
                                ft.Text(label, size=9, weight=ft.FontWeight.BOLD,
                                        color=colors["subtext"], text_align=ft.TextAlign.CENTER),
                                val_lbl,
                            ],
                            horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=4,
                        ),
                        padding=ft.padding.only(top=10, bottom=12),
                        alignment=ft.alignment.center, expand=True,
                    ),
                ],
                spacing=0,
            ),
            bgcolor=colors["surface"], border_radius=8, padding=0, expand=True,
            border=ft.border.all(1, colors["border"]),
        )
        card.val_lbl = val_lbl
        return card

    def _build_stats(self, colors: dict):
        c_total = colors["text"] if not self.is_dark else "#7fb3e0"
        cards = [
            ("TOTAL REGISTROS", str(self.total), c_total),
            ("PROCESADOS OK",   str(self.processed), C_STAT_OK),
            ("ERRORES",         str(self.errors), C_STAT_ERR),
            ("PENDIENTES",      str(max(0, self.total - self.processed - self.errors)), C_STAT_PEN),
        ]
        self.stat_vals: list[ft.Text] = []
        cells = []
        for label, value, color in cards:
            card = self._stat_card(colors, label, value, color)
            self.stat_vals.append(card.val_lbl)
            cells.append(card)
        self.stats_row = ft.Container(ft.Row(cells, spacing=10), padding=ft.padding.symmetric(horizontal=12))

    def _build_progress(self, colors: dict):
        self.status_lbl = ft.Text(self.status_text, color=colors["subtext"], size=11, expand=True)
        self.elapsed_lbl = ft.Text(self.elapsed_text, color=colors["subtext"], size=11)
        self.progress_bar = ft.ProgressBar(value=self.progress_val, height=14, expand=True,
                                            color=C_GOLD, bgcolor=colors["border"])
        self.pct_lbl = ft.Text(f"{int(self.progress_val * 100)}%", width=42,
                                weight=ft.FontWeight.BOLD, size=11, color=colors["text"])

        self.progress_col = ft.Container(
            ft.Column(
                [
                    ft.Row([self.status_lbl, self.elapsed_lbl]),
                    ft.Row([self.progress_bar, self.pct_lbl]),
                ],
                spacing=4,
            ),
            padding=ft.padding.symmetric(horizontal=12),
        )

    def _build_logs(self, colors: dict):
        log_text_color = LOG_COLOR_DARK if self.is_dark else LOG_COLOR_LIGHT
        self.log_list = ft.ListView(spacing=2, auto_scroll=True, expand=True)
        for text, level in self.log_entries:
            self.log_list.controls.append(
                ft.Text(text, font_family="Courier New", size=11,
                        color=log_text_color.get(level, colors["text"]), selectable=True)
            )
        self.log_card = ft.Container(
            ft.Column(
                [
                    ft.Row(
                        [
                            ft.Text("Registro de actividad", weight=ft.FontWeight.BOLD, color=colors["text"]),
                            ft.Container(expand=True),
                            ft.TextButton("Limpiar", on_click=self._clear_logs),
                        ],
                    ),
                    ft.Container(self.log_list, expand=True, bgcolor=colors["log_bg"],
                                border_radius=6, padding=8,
                                border=ft.border.all(1, colors["border"])),
                ],
                expand=True,
            ),
            bgcolor=colors["surface"], border_radius=10, padding=12, expand=True, height=320,
            border=ft.border.all(1, colors["border"]),
        )

    def _build_buttons(self, colors: dict):
        self.btn_run = ft.ElevatedButton(
            "EJECUTAR RPA", icon=ft.Icons.PLAY_ARROW,
            on_click=self._start, disabled=self.is_running,
            style=ft.ButtonStyle(bgcolor=C_BTN_RUN, color="white"),
            height=44,
        )
        self.btn_cancel = ft.ElevatedButton(
            "CANCELAR", icon=ft.Icons.STOP,
            on_click=self._cancel, disabled=not self.is_running,
            style=ft.ButtonStyle(bgcolor=C_BTN_STOP, color="white"),
            height=44,
        )
        self.btn_compare = ft.ElevatedButton(
            "COMPARAR", icon=ft.Icons.BAR_CHART,
            on_click=self._start_compare,
            style=ft.ButtonStyle(bgcolor=C_BTN_CMP, color="white"),
            height=44,
        )
        self.auto_compare_chk = ft.Switch(
            label="Auto-comparar al terminar", value=self.auto_compare,
            label_position=ft.LabelPosition.RIGHT,
            on_change=lambda e: setattr(self, "auto_compare", e.control.value),
        )
        self.btn_catalogs = ft.ElevatedButton(
            "CATÁLOGOS", icon=ft.Icons.FOLDER_COPY,
            on_click=lambda e: self.page.go("/catalogos"),
            style=ft.ButtonStyle(bgcolor=C_BTN_CAT, color="white"),
            height=44,
        )
        self.btn_poliza = ft.ElevatedButton(
            "PÓLIZA TXT", icon=ft.Icons.RECEIPT_LONG,
            on_click=self._start_poliza,
            style=ft.ButtonStyle(bgcolor=C_BTN_POL, color="white"),
            height=44,
        )

        self.buttons_row = ft.Row(
            [
                self.btn_run, self.btn_cancel,
                ft.VerticalDivider(width=1),
                self.btn_compare, self.auto_compare_chk,
                ft.VerticalDivider(width=1),
                self.btn_catalogs,
                ft.VerticalDivider(width=1),
                self.btn_poliza,
            ],
            wrap=False, spacing=10, scroll=ft.ScrollMode.AUTO,
        )

    # ────────────────────────────────────────────────────
    # Logging
    # ────────────────────────────────────────────────────
    def _log(self, msg: str, level: str = "info"):
        ts = datetime.now().strftime("%H:%M:%S")
        icon = LOG_ICON.get(level, "·")
        text = f"[{ts}] {icon}  {msg}"
        self.log_entries.append((text, level))
        log_text_color = LOG_COLOR_DARK if self.is_dark else LOG_COLOR_LIGHT
        self.log_list.controls.append(
            ft.Text(text, font_family="Courier New", size=11,
                    color=log_text_color.get(level, self.colors["text"]), selectable=True)
        )
        self.page.update()

    def _clear_logs(self, e):
        self.log_entries.clear()
        self.log_list.controls.clear()
        self.page.update()

    # ────────────────────────────────────────────────────
    # File handling
    # ────────────────────────────────────────────────────
    def _browse_file(self, e):
        self.file_picker.pick_files(
            dialog_title="Seleccionar Excel de Recepción de Facturas",
            initial_directory=str(RECEPCION_DIR) if RECEPCION_DIR.exists() else None,
            allowed_extensions=["xlsx", "xls"],
        )

    def _on_file_picked(self, e: ft.FilePickerResultEvent):
        if not e.files:
            return
        path = e.files[0].path
        self.file_path = path
        self.file_field.value = path
        self.page.update()
        self._load_file_info(path)

    def _load_file_info(self, path: str):
        try:
            from rpa.excel_handler import ExcelHandler
            handler = ExcelHandler(path)
            folios = handler.get_folios()
            n = len(folios)
            self.total = n
            self.stat_vals[0].value = str(n)
            self.stat_vals[3].value = str(n)
            self.status_text = f"Listo — {n} folio(s) en '{Path(path).name}'"
            self.status_lbl.value = self.status_text
            self.page.update()
            self._log(f"Archivo cargado: {Path(path).name}  ({n} folios)", "info")
        except Exception as exc:
            self._log(f"Error leyendo Excel: {exc}", "error")

    # ────────────────────────────────────────────────────
    # Progress helpers
    # ────────────────────────────────────────────────────
    def _refresh_stats(self, processed: int, errors: int, folio: str = ""):
        self.processed = processed
        self.errors = errors
        pending = max(0, self.total - processed - errors)
        self.stat_vals[1].value = str(processed)
        self.stat_vals[2].value = str(errors)
        self.stat_vals[3].value = str(pending)
        if self.total > 0:
            pct = (processed + errors) / self.total
            self.progress_val = pct
            self.progress_bar.value = pct
            self.pct_lbl.value = f"{int(pct * 100)}%"
        if folio:
            self.status_text = f"Procesando: {folio}"
            self.status_lbl.value = self.status_text
        self.page.update()

    async def _tick(self):
        while self.is_running and self.start_ts:
            secs = int((datetime.now() - self.start_ts).total_seconds())
            m, s = divmod(secs, 60)
            self.elapsed_text = f"Tiempo: {m:02d}:{s:02d}"
            self.elapsed_lbl.value = self.elapsed_text
            self.page.update()
            await asyncio.sleep(1)

    # ────────────────────────────────────────────────────
    # RPA lifecycle
    # ────────────────────────────────────────────────────
    def _start(self, e):
        if not self.file_path:
            self.file_field.error_text = "Selecciona un archivo Excel"
            self.page.update()
            return
        self.file_field.error_text = None
        if not self.user_field.value or not self.pass_field.value:
            if not self.user_field.value:
                self.user_field.error_text = "Requerido"
            if not self.pass_field.value:
                self.pass_field.error_text = "Requerido"
            self.page.update()
            return
        self.user_field.error_text = None
        self.pass_field.error_text = None

        self.is_running = True
        self.processed = 0
        self.errors = 0
        self.start_ts = datetime.now()

        for lbl in self.stat_vals[1:]:
            lbl.value = "0"
        self.progress_val = 0
        self.progress_bar.value = 0
        self.pct_lbl.value = "0%"
        self.btn_run.disabled = True
        self.btn_cancel.disabled = False
        self.status_text = "Iniciando proceso..."
        self.status_lbl.value = self.status_text
        self.page.update()

        self._log("─" * 62, "info")
        self._log(
            f"Inicio: {self.start_ts.strftime('%d/%m/%Y %H:%M:%S')}  |  "
            f"Archivo: {Path(self.file_path).name}",
            "info",
        )

        self.page.run_task(self._tick)
        self.page.run_task(self._async_run)

    def _cancel(self, e):
        if self.is_running:
            self.is_running = False
            self._log("Cancelación solicitada — terminando folio actual...", "warn")
            self.status_text = "Cancelando..."
            self.status_lbl.value = self.status_text
            self.page.update()

    async def _async_run(self):
        from rpa.automation import RPAAutomation
        from rpa.excel_handler import ExcelHandler

        try:
            file_path = self.file_path
            auto_compare = self.auto_compare_chk.value
            handler = ExcelHandler(file_path)
            folio_rows = handler.get_folios()
            if MAX_RECORDS is not None:
                folio_rows = folio_rows[:MAX_RECORDS]
                self._log(f"[--max {MAX_RECORDS}] Limitando a los primeros {len(folio_rows)} folio(s).", "warn")
            self.total = len(folio_rows)
            self.stat_vals[0].value = str(self.total)
            self.page.update()

            n_workers = int(self.workers_selected)
            n_workers = min(n_workers, self.total) or 1

            self._log(
                f"Iniciando RPA — {self.total} folio(s) en {n_workers} sesión(es) paralela(s)…",
                "info",
            )

            chunk_sz = (self.total + n_workers - 1) // n_workers
            chunks = [
                folio_rows[i * chunk_sz:(i + 1) * chunk_sz]
                for i in range(n_workers)
                if folio_rows[i * chunk_sz:(i + 1) * chunk_sz]
            ]
            actual_workers = len(chunks)

            proc_counts = [0] * actual_workers
            err_counts = [0] * actual_workers

            def make_progress_cb(idx: int):
                def _cb(processed: int, errors: int, folio: str):
                    proc_counts[idx] = processed
                    err_counts[idx] = errors
                    label = f"W{idx+1}›{folio}" if actual_workers > 1 else folio
                    self._refresh_stats(sum(proc_counts), sum(err_counts), label)
                return _cb

            async def run_worker(idx: int, chunk):
                prefix = f"[W{idx+1}] " if actual_workers > 1 else ""

                def _log(msg: str, level: str = "info"):
                    self._log(f"{prefix}{msg}", level)

                rpa = RPAAutomation(
                    username=self.username,
                    password=self.password,
                    headless=self.headless,
                    log_fn=_log,
                    cancel_fn=lambda: not self.is_running,
                )
                worker_results = await rpa.run(chunk, on_progress=make_progress_cb(idx))
                return worker_results, rpa

            all_returns = await asyncio.gather(
                *[run_worker(i, chunk) for i, chunk in enumerate(chunks)],
                return_exceptions=True,
            )

            results: list = []
            all_skipped: list = []
            all_not_found: list = []

            for i, ret in enumerate(all_returns):
                if isinstance(ret, BaseException):
                    import traceback as _tb
                    self._log(f"[W{i+1}] Error fatal: {ret}", "error")
                    self._log(_tb.format_exception(type(ret), ret, ret.__traceback__)[-1], "error")
                else:
                    w_results, rpa = ret
                    results.extend(w_results)
                    all_skipped.extend(getattr(rpa, "skipped", []))
                    all_not_found.extend(getattr(rpa, "not_found", []))

            results.sort(key=lambda r: r[0])

            if results:
                self._log("Guardando resultados en Excel…", "info")
                handler.ensure_headers()
                for row_num, cc, obs, subtotal, descuento, iva, gastos_envio, total_oc, cuentas_contables, poliza_lineas in results:
                    handler.write_result(row_num, cc, obs, subtotal, descuento, iva, gastos_envio, total_oc, cuentas_contables, poliza_lineas)
                handler.save()
                self._log(f"Excel actualizado: {Path(file_path).name}", "ok")

                if auto_compare:
                    await self._run_compare(file_path)

            ok_count = sum(1 for r in results if r[1])
            elapsed = int((datetime.now() - self.start_ts).total_seconds())
            m, s = divmod(elapsed, 60)

            self._log("═" * 62, "info")
            self._log("RESUMEN FINAL", "info")
            self._log(f"  Total folios:        {self.total}", "info")
            self._log(f"  Sesiones usadas:     {actual_workers}", "info")
            self._log(f"  Con CC extraído:     {ok_count}", "ok")
            self._log(f"  Sin OC / con error:  {self.total - ok_count}", "warn")
            if all_skipped:
                self._log(f"  Duplicados omitidos: {len(all_skipped)}", "warn")
                for f in all_skipped:
                    self._log(f"    · {f}", "warn")
            if all_not_found:
                self._log(f"  No encontrados:      {len(all_not_found)}", "warn")
                for f in all_not_found:
                    self._log(f"    · {f}", "warn")
            self._log(f"  Tiempo total:        {m}m {s}s", "info")
            self._log("═" * 62, "info")

            self.status_text = f"Completado — {ok_count}/{self.total} con CC extraído  ({m}m {s}s)"
            self.status_lbl.value = self.status_text
            self.page.update()

        except Exception as exc:
            import traceback
            self._log(f"Error crítico: {exc}", "error")
            self._log(traceback.format_exc(), "error")

        finally:
            self.is_running = False
            self.btn_run.disabled = False
            self.btn_cancel.disabled = True
            self.page.update()

    # ────────────────────────────────────────────────────
    # Comparar sucursales
    # ────────────────────────────────────────────────────
    def _start_compare(self, e):
        self.compare_picker.pick_files(
            dialog_title="Seleccionar archivo para comparar sucursales",
            initial_directory=str(RECEPCION_DIR) if RECEPCION_DIR.exists() else None,
            allowed_extensions=["xlsx", "xls", "csv"],
        )

    def _on_compare_picked(self, e: ft.FilePickerResultEvent):
        if not e.files:
            return
        path = e.files[0].path
        self.page.run_task(self._run_compare, path)

    async def _run_compare(self, path: str):
        self._log("─" * 62, "info")
        self._log(f"Comparando sucursales: {Path(path).name}", "info")
        self.status_text = f"Comparando: {Path(path).name}…"
        self.status_lbl.value = self.status_text
        self.btn_compare.disabled = True
        self.page.update()

        try:
            await asyncio.to_thread(self._compare_blocking, path)
        finally:
            self.btn_compare.disabled = False
            self.page.update()

    def _compare_blocking(self, csv_path: str):
        try:
            import compare_sucursales as cs
            from openpyxl import Workbook

            self._log("  Cargando registros…", "info")
            header, data = cs.load_file(csv_path)
            total = len(data)
            self._log(f"  {total} registros cargados.", "info")

            catalog = cs.load_catalogs(cs.DISTRIBUCION_DIR)
            ut_catalog = cs.load_utilitario_catalogs(cs.DISTRIBUCION_DIR)
            all_cuentas = cs.load_all_cuentas()
            if catalog:
                self._log(f"  {len(catalog)} claves de catálogo de distribución cargadas.", "info")
            else:
                self._log("  [AVISO] Carpeta Distribucion/ no encontrada — sin cálculo de montos.", "warn")
            if ut_catalog:
                self._log(f"  {len(ut_catalog)} utilitarios cargados.", "info")
            if all_cuentas:
                self._log(f"  {len(all_cuentas)} cuentas contables cargadas (gastos + proveedores).", "info")
            else:
                self._log("  [AVISO] Catálogos de cuentas no encontrados — sin conciliación.", "warn")

            wb = Workbook()
            ws_orig = wb.active
            ws_sum = wb.create_sheet()
            ws_sum_cc = wb.create_sheet()
            ws_suc = wb.create_sheet()
            ws_main = wb.create_sheet()
            ws_dist = wb.create_sheet()
            ws_dist_calc = wb.create_sheet()

            counts, details = cs.build_main_sheet(ws_main, data, ut_catalog)
            cc_counts = cs.build_datos_originales_sheet(ws_orig, header, data, all_cuentas)
            cs.build_summary_sheet(ws_sum, counts, total)
            cs.build_resumen_cuentas_sheet(ws_sum_cc, cc_counts, total)
            cs.build_sucursal_detail_sheet(ws_suc, details)
            cs.build_distribucion_sheet(ws_dist, details)
            cs.build_distribucion_calculada_sheet(ws_dist_calc, details, catalog)

            out = Path(csv_path).parent / (Path(csv_path).stem + "_comparacion.xlsx")
            wb.save(str(out))

            match_pct = counts["MATCH"] / total * 100 if total else 0
            mis_pct = counts["MISMATCH"] / total * 100 if total else 0
            dis_pct = counts["DISTRIBUCIÓN"] / total * 100 if total else 0
            sin_pct = counts["SIN SUCURSAL"] / total * 100 if total else 0

            self._log("═" * 62, "info")
            self._log("RESUMEN COMPARACIÓN DE SUCURSALES", "info")
            self._log(f"  Total registros  : {total}", "info")
            self._log(f"  MATCH ✓          : {counts['MATCH']}  ({match_pct:.1f}%)", "ok")
            self._log(f"  MISMATCH ✗       : {counts['MISMATCH']}  ({mis_pct:.1f}%)", "error")
            self._log(f"  DISTRIBUCIÓN     : {counts['DISTRIBUCIÓN']}  ({dis_pct:.1f}%)", "warn")
            self._log(f"  Sin sucursal     : {counts['SIN SUCURSAL']}  ({sin_pct:.1f}%)", "info")
            self._log("─" * 62, "info")
            self._log("RESUMEN CUENTAS CONTABLES", "info")
            self._log(f"  Match ✓          : {cc_counts['match']}", "ok")
            self._log(f"  Mismatch ✗       : {cc_counts['mismatch']}", "error")
            self._log(f"  Cuadre           : {cc_counts.get('cuadre', 0)}", "warn")
            self._log(f"  Sin dato SIPP    : {cc_counts['sin_dato_sipp']}", "warn")
            self._log(f"  Archivo generado : {out.name}", "ok")
            self._log("═" * 62, "info")

            self.status_text = (
                f"Comparación lista — {counts['MATCH']} MATCH  |  "
                f"{counts['MISMATCH']} MISMATCH  |  "
                f"{counts['DISTRIBUCIÓN']} DISTRIBUCIÓN"
            )
            self.status_lbl.value = self.status_text
            self.page.update()

        except Exception as exc:
            import traceback
            self._log(f"Error en comparación: {exc}", "error")
            self._log(traceback.format_exc(), "error")
            self.status_text = "Error en comparación."
            self.status_lbl.value = self.status_text
            self.page.update()

    # ────────────────────────────────────────────────────
    # Generar póliza TXT (Contpaq)
    # ────────────────────────────────────────────────────
    def _start_poliza(self, e):
        self._poliza_num_field = ft.TextField(
            label="Número de póliza (campo NumPol en Contpaq)",
            value="1", keyboard_type=ft.KeyboardType.NUMBER, autofocus=True,
        )
        self._poliza_esquema_group = ft.RadioGroup(
            value="sipp",
            content=ft.Column(
                [
                    ft.Radio(value="sipp", label="SIPP — Provisión/Entrada/Salida (cuentas globales)"),
                    ft.Radio(value="individual", label="Individual — 1 póliza, cuentas por estación + proveedor real"),
                ],
                spacing=2,
            ),
        )

        def _continue(ev):
            try:
                num_poliza = int(self._poliza_num_field.value)
                if num_poliza < 1:
                    raise ValueError
            except ValueError:
                self._poliza_num_field.error_text = "Ingresa un número entero ≥ 1"
                self.page.update()
                return
            dlg.open = False
            self.page.update()
            self._poliza_num = num_poliza
            self._poliza_esquema = self._poliza_esquema_group.value
            # En desktop, abrir el panel nativo de archivos en el mismo instante en
            # que se cierra este diálogo modal puede hacer que el SO no lo muestre —
            # se da un respiro async para que el cierre termine de pintarse primero.
            self.page.run_task(self._open_poliza_picker)

        def _cancel(ev):
            dlg.open = False
            self.page.update()

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("Generar póliza"),
            content=ft.Column(
                [self._poliza_num_field, ft.Divider(height=1), self._poliza_esquema_group],
                tight=True, width=420, spacing=10,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=_cancel),
                ft.FilledButton("Continuar", on_click=_continue),
            ],
        )
        self.page.open(dlg)

    async def _open_poliza_picker(self):
        await asyncio.sleep(0.3)
        try:
            self.poliza_picker.pick_files(
                dialog_title="Seleccionar reporte *_comparacion.xlsx",
                initial_directory=str(RECEPCION_DIR) if RECEPCION_DIR.exists() else None,
                allowed_extensions=["xlsx"],
            )
        except Exception as exc:
            self._log(f"No se pudo abrir el selector de archivo: {exc}", "error")

    def _on_poliza_picked(self, e: ft.FilePickerResultEvent):
        if not e.files:
            return
        path = e.files[0].path
        num_poliza = getattr(self, "_poliza_num", 1)
        esquema = getattr(self, "_poliza_esquema", "sipp")

        self._log("─" * 62, "info")
        self._log(f"Generando póliza TXT ({esquema}) desde: {Path(path).name}", "info")
        self.status_text = f"Generando póliza: {Path(path).name}…"
        self.status_lbl.value = self.status_text
        self.btn_poliza.disabled = True
        self.page.update()

        self.page.run_task(self._poliza_task, path, num_poliza, esquema)

    async def _poliza_task(self, xlsx_path: str, num_poliza: int, esquema: str):
        try:
            await asyncio.to_thread(self._poliza_blocking, xlsx_path, num_poliza, esquema)
        finally:
            self.btn_poliza.disabled = False
            self.page.update()

    def _poliza_blocking(self, xlsx_path: str, num_poliza: int, esquema: str):
        try:
            if esquema == "individual":
                self._poliza_blocking_individual(xlsx_path, num_poliza)
            else:
                self._poliza_blocking_sipp(xlsx_path, num_poliza)
        except Exception as exc:
            import traceback
            self._log(f"Error generando póliza: {exc}", "error")
            self._log(traceback.format_exc(), "error")
            self.status_text = "Error generando póliza."
            self.status_lbl.value = self.status_text
            self.page.update()

    def _poliza_blocking_sipp(self, xlsx_path: str, num_poliza: int):
        from rpa.poliza_generator import generar_polizas_almacen

        resultado = generar_polizas_almacen(
            xlsx_path,
            num_poliza_provision=num_poliza,
            num_poliza_entrada=num_poliza,
            num_poliza_salida=num_poliza,
            log_fn=self._log,
        )

        self._log("═" * 62, "info")
        self._log("RESUMEN PÓLIZAS (SIPP)", "info")
        self._log(f"  Facturas procesadas      : {resultado['procesadas']}", "ok")
        self._log(f"  Sin cuenta de almacén    : {len(resultado['sin_almacen'])}", "warn")
        self._log(f"  Sin cuenta contable SIPP : {len(resultado['sin_cuenta_gasto'])}", "warn")
        self._log("═" * 62, "info")

        self.status_text = (
            f"Pólizas generadas — {resultado['procesadas']} factura(s)  |  "
            f"{len(resultado['sin_almacen'])} sin almacén"
        )
        self.status_lbl.value = self.status_text
        self.page.update()

        if resultado.get("sucursales_sin_almacen"):
            self._show_missing_almacen_dialog(resultado["sucursales_sin_almacen"])

    def _poliza_blocking_individual(self, xlsx_path: str, num_poliza: int):
        from rpa.poliza_generator import generar_poliza_individual

        resultado = generar_poliza_individual(
            xlsx_path,
            num_poliza=num_poliza,
            log_fn=self._log,
        )

        self._log("═" * 62, "info")
        self._log("RESUMEN PÓLIZA INDIVIDUAL", "info")
        self._log(f"  Facturas procesadas      : {resultado['procesadas']}", "ok")
        self._log(f"  Archivos generados       : {len(resultado['archivos'])} (uno por fecha)", "ok")
        self._log(f"  Comprimido (.zip)        : {Path(resultado['zip_path']).name}", "ok")
        self._log(f"  Sin póliza SIPP real     : {len(resultado['sin_poliza_sipp'])}", "warn")
        self._log(f"  Sin cuenta de estación   : {len(resultado['sin_estacion'])}", "warn")
        self._log(f"  Sin cuenta de proveedor  : {len(resultado['sin_proveedor'])}", "warn")
        self._log("═" * 62, "info")

        self.status_text = (
            f"Póliza Individual generada — {resultado['procesadas']} factura(s)  |  "
            f"{len(resultado['sin_estacion']) + len(resultado['sin_proveedor']) + len(resultado['sin_poliza_sipp'])} omitida(s)"
        )
        self.status_lbl.value = self.status_text
        self.page.update()

        if resultado.get("estaciones_sin_cuenta") or resultado.get("proveedores_sin_cuenta"):
            self._show_missing_individual_dialog(
                resultado.get("estaciones_sin_cuenta") or [],
                resultado.get("proveedores_sin_cuenta") or [],
            )

    def _show_missing_almacen_dialog(self, sucursales: list[str]):
        """Resumen de sucursales sin cuenta de Almacén Ref — permite agregarlas
        al catálogo con un clic, dejando solo la celda 'Cuenta' vacía por llenar."""
        items = ft.Column(
            [ft.Text(f"•  {s}", size=12) for s in sucursales],
            tight=True, spacing=2, scroll=ft.ScrollMode.AUTO,
            height=min(220, 22 * len(sucursales) + 10),
        )

        def _dismiss(e):
            dlg.open = False
            self.page.update()

        def _add_and_open(e):
            dlg.open = False
            self.pending_almacen_rows = sucursales
            self.catalogs_initial_tab = next(
                (i for i, spec in enumerate(CATALOG_TABS) if spec["name"] == "Almacén Ref"), 0
            )
            self.page.update()
            self.page.go("/catalogos")

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("Sucursales sin cuenta de almacén"),
            content=ft.Column(
                [
                    ft.Text(
                        f"{len(sucursales)} sucursal(es) de esta póliza no tienen cuenta de "
                        "Almacén Ref configurada. Se pueden agregar al catálogo dejando la "
                        "celda 'Cuenta' vacía — solo tendrías que capturar el código.",
                        size=12,
                    ),
                    ft.Divider(height=1),
                    items,
                ],
                tight=True, width=380, spacing=8,
            ),
            actions=[
                ft.TextButton("Más tarde", on_click=_dismiss),
                ft.FilledButton(
                    "Agregar a Catálogo", icon=ft.Icons.ADD, on_click=_add_and_open,
                    style=ft.ButtonStyle(bgcolor=C_BTN_CAT, color="white"),
                ),
            ],
        )
        self.page.open(dlg)

    def _show_missing_individual_dialog(self, estaciones: list[str], proveedores: list[str]):
        """Resumen de facturas omitidas de la Póliza Individual por falta de
        cuenta de estación o de proveedor en los catálogos."""
        sections: list[ft.Control] = []

        if estaciones:
            sections.append(ft.Text(
                f"{len(estaciones)} estación(es) sin cuenta en 'Cuentas Gasto Estaciones' "
                "(facturas omitidas). No se prellenan filas — cada estación nueva necesita "
                "varias cuentas — agrégalas manualmente en esa pestaña:",
                size=12,
            ))
            sections.append(ft.Column(
                [ft.Text(f"•  {s}", size=12) for s in estaciones],
                tight=True, spacing=2, scroll=ft.ScrollMode.AUTO,
                height=min(140, 22 * len(estaciones) + 10),
            ))
            sections.append(ft.Divider(height=1))

        if proveedores:
            sections.append(ft.Text(
                f"{len(proveedores)} proveedor(es) sin cuenta en 'Cuentas Proveedores' "
                "(facturas omitidas). Se pueden agregar al catálogo dejando la celda "
                "'Cuenta' vacía:",
                size=12,
            ))
            sections.append(ft.Column(
                [ft.Text(f"•  {p}", size=12) for p in proveedores],
                tight=True, spacing=2, scroll=ft.ScrollMode.AUTO,
                height=min(140, 22 * len(proveedores) + 10),
            ))

        def _dismiss(e):
            dlg.open = False
            self.page.update()

        def _ver_estaciones(e):
            dlg.open = False
            self.catalogs_initial_tab = next(
                (i for i, spec in enumerate(CATALOG_TABS) if spec["name"] == "Cuentas Gasto Estaciones"), 0
            )
            self.page.update()
            self.page.go("/catalogos")

        def _agregar_proveedores(e):
            dlg.open = False
            self.pending_proveedor_rows = proveedores
            self.catalogs_initial_tab = next(
                (i for i, spec in enumerate(CATALOG_TABS) if spec["name"] == "Cuentas Proveedores"), 0
            )
            self.page.update()
            self.page.go("/catalogos")

        actions = [ft.TextButton("Cerrar", on_click=_dismiss)]
        if estaciones:
            actions.append(ft.OutlinedButton("Ver Cuentas Gasto Estaciones", on_click=_ver_estaciones))
        if proveedores:
            actions.append(ft.FilledButton(
                "Agregar proveedores faltantes", icon=ft.Icons.ADD, on_click=_agregar_proveedores,
                style=ft.ButtonStyle(bgcolor=C_BTN_CAT, color="white"),
            ))

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("Facturas omitidas — Póliza Individual"),
            content=ft.Column(sections, tight=True, width=420, spacing=8),
            actions=actions,
        )
        self.page.open(dlg)

    # ────────────────────────────────────────────────────
    # Auto-update
    # ────────────────────────────────────────────────────
    @staticmethod
    def _git(args: list, cwd: str, timeout: int = 10) -> str:
        kwargs = dict(cwd=cwd, capture_output=True, text=True, encoding="utf-8",
                      errors="replace", timeout=timeout)
        if sys.platform == "win32":
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        r = subprocess.run(["git"] + args, **kwargs)
        return r.stdout.strip() if r.returncode == 0 else ""

    def _check_updates_sync(self, *, manual: bool = False):
        root = str(ROOT_DIR)
        try:
            fkwargs = dict(cwd=root, capture_output=True, text=True,
                          encoding="utf-8", errors="replace", timeout=20)
            if sys.platform == "win32":
                fkwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
            fetch = subprocess.run(["git", "fetch", "origin", "main", "--quiet"], **fkwargs)
            if fetch.returncode != 0:
                if manual:
                    self._log("No se pudo conectar a GitHub (sin git o sin red).", "warn")
                return

            local = self._git(["rev-parse", "HEAD"], cwd=root)
            remote = self._git(["rev-parse", "origin/main"], cwd=root)
            if not local or not remote:
                return
            if local == remote:
                if manual:
                    self._log("La app está al día. ✓", "ok")
                return

            behind = self._git(["rev-list", "--count", "HEAD..origin/main"], cwd=root)
            last_msg = self._git(["log", "-1", "--pretty=%s", "origin/main"], cwd=root)
            self._show_update_banner(behind or "?", last_msg or "")
        except Exception:
            pass

    def _show_update_banner(self, behind: str, last_msg: str):
        self.update_banner_text = f"  🔄  {behind} actualización(es) disponible(s)  ·  \"{last_msg}\""
        self.update_banner_visible = True
        self.update_lbl.value = self.update_banner_text
        self.update_banner.visible = True
        self.page.update()
        self._log(f"Actualización disponible ({behind} commit(s)): {last_msg}", "warn")

    def _dismiss_update_banner(self, e):
        self.update_banner_visible = False
        self.update_banner.visible = False
        self.page.update()

    def _do_update(self, e):
        self.btn_update.disabled = True
        self.btn_update.text = "Actualizando…"
        self.page.update()
        threading.Thread(target=self._pull_and_restart, daemon=True).start()

    def _pull_and_restart(self):
        root = str(ROOT_DIR)
        try:
            res = subprocess.run(["git", "pull", "origin", "main"], cwd=root,
                                 capture_output=True, text=True, timeout=60)
            if res.returncode == 0:
                self._restart_app()
            else:
                err = res.stderr.strip() or res.stdout.strip()
                self._log(f"git pull falló: {err}", "error")
                self.btn_update.disabled = False
                self.btn_update.text = "Actualizar y reiniciar"
                self.page.update()
        except Exception as exc:
            self._log(f"Error al actualizar: {exc}", "error")

    def _restart_app(self):
        if sys.platform == "win32":
            subprocess.Popen([sys.executable] + sys.argv)
            sys.exit(0)
        else:
            import os
            os.execv(sys.executable, [sys.executable] + sys.argv)

    def _manual_check_updates(self, e):
        self.btn_check_update.disabled = True
        self.page.update()
        self._log("Comprobando actualizaciones…", "info")

        def _check_and_restore():
            self._check_updates_sync(manual=True)
            self.btn_check_update.disabled = False
            self.page.update()

        threading.Thread(target=_check_and_restore, daemon=True).start()


def main(page: ft.Page):
    RPAApp(page)


if __name__ == "__main__":
    ft.app(target=main)
